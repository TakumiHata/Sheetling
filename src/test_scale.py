import sys
from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
for col_idx in range(1, 120 + 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 2.71
for row_idx in range(1, 170 + 1):
    ws.row_dimensions[row_idx].height = 18.0

ws.sheet_view.zoomScale = 35
ws.page_margins = PageMargins(left=0.2, right=0.2, top=0.4, bottom=0.4, header=0, footer=0)
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.paperSize = 9
ws.page_setup.orientation = "portrait"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.print_options.horizontalCentered = True
ws.print_area = "A1:DP170"

ws.cell(row=170, column=120, value="TEST")
wb.save("data/04_prompt/test_1x1.xlsx")
print("Saved data/04_prompt/test_1x1.xlsx")
