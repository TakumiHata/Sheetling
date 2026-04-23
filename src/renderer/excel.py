import re

from src.templates.prompts import GRID_SIZES
from src.utils.logger import get_logger

logger = get_logger(__name__)


# キャッシュされた _grid_params.json から読み込んだ場合でも、
# 描画系パラメータ（列幅・行高・フォント・余白）はコード変更を反映するため
# レンダー時に GRID_SIZES から再解決する。
_RENDER_ONLY_KEYS = (
    'excel_col_width', 'excel_row_height',
    'default_font_size', 'font_name',
    'margin_left', 'margin_right', 'margin_top', 'margin_bottom',
)


def _refresh_render_params(grid_params: dict) -> dict:
    grid_size = grid_params.get('grid_size', '1pt')
    is_a3 = grid_params.get('paper_size') == 8
    ref_key = f"{grid_size}_a3" if is_a3 else grid_size
    ref = GRID_SIZES.get(ref_key, GRID_SIZES.get(grid_size, GRID_SIZES["1pt"]))
    refreshed = dict(grid_params)
    for key in _RENDER_ONLY_KEYS:
        if key in ref:
            refreshed[key] = ref[key]
    return refreshed


def fix_empty_cell_type_attr(xlsx_path: str) -> None:
    """
    openpyxl 3.1.x が空セルに付与する t="n" 属性を除去する。
    Excel Online がこの属性を不正として罫線スタイルを除去してしまう問題の回避策。
    """
    import zipfile, shutil
    pat = re.compile(r'(<c\s+r="[^"]+"\s+s="\d+"\s+)t="n"\s*(/>)')
    tmp = xlsx_path + '.tmp_fix'
    with zipfile.ZipFile(xlsx_path, 'r') as zin, zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                text = data.decode('utf-8')
                text = pat.sub(r'\1\2', text)
                data = text.encode('utf-8')
            zout.writestr(item, data)
    shutil.move(tmp, xlsx_path)


def _create_workbook(grid_params: dict):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    col_width = grid_params.get('excel_col_width', 1.45)
    ws.sheet_format.defaultColWidth = col_width
    ws.sheet_format.defaultRowHeight = grid_params.get('excel_row_height', 11.34)
    ws.sheet_format.customHeight = True
    ws.sheet_view.showGridLines = True
    # defaultColWidth だけでは Excel が独自換算で狭く表示するため、
    # 印刷範囲をカバーする全列に明示的な width を設定する。
    max_cols = grid_params.get('max_cols', 50)
    for col_idx in range(1, max_cols + 2):  # +1 for COL_OFFSET (左1列空け)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    return wb, ws


def _place_text_element(ws, elem, row_offset, col_offset, default_font_size, font_name):
    from openpyxl.styles import Alignment, Font
    r = elem.get('row', 1) + row_offset
    c = elem.get('col', 1) + col_offset
    try:
        cell = ws.cell(row=r, column=c)
        cell.value = elem.get('content', '')
        if elem.get('is_vertical'):
            cell.alignment = Alignment(text_rotation=255, vertical='top', wrap_text=False)
        else:
            cell.alignment = Alignment(
                horizontal='left', vertical='top',
                wrap_text=bool(elem.get('multiline')),
            )
        resolved_name = elem.get('font_name') or font_name
        resolved_size = float(elem.get('font_size') or default_font_size)
        font_kwargs = {'name': resolved_name, 'size': resolved_size}
        if elem.get('font_color'):
            font_kwargs['color'] = elem['font_color']
        cell.font = Font(**font_kwargs)
    except AttributeError:
        pass
    return r, c


def _place_border_element(ws, elem, row_offset, col_offset):
    from openpyxl.styles import Border, Side
    s_row = elem.get('row', 1) + row_offset
    e_row = elem.get('end_row', 1) + row_offset
    s_col = elem.get('col', 1) + col_offset
    e_col = elem.get('end_col', 1) + col_offset
    borders = elem.get('borders', {'top': True, 'bottom': True, 'left': True, 'right': True})
    border_style = elem.get('border_style', 'thin')
    side = Side(style=border_style)

    def set_side(row, col, **sides):
        if row < 1:
            return
        try:
            cell = ws.cell(row=row, column=col)
            ex = cell.border
            cell.border = Border(
                top=sides.get('top', ex.top), bottom=sides.get('bottom', ex.bottom),
                left=sides.get('left', ex.left), right=sides.get('right', ex.right),
            )
        except AttributeError:
            pass

    if borders.get('top', True):
        for c in range(s_col, e_col):
            set_side(s_row, c, top=side)
    if borders.get('bottom', True):
        for c in range(s_col, e_col):
            set_side(e_row - 1, c, bottom=side)
    if borders.get('left', True):
        for r in range(s_row, e_row):
            set_side(r, s_col, left=side)
    if borders.get('right', True):
        for r in range(s_row, e_row):
            set_side(r, e_col - 1, right=side)
    return e_row, e_col


def _finalize_workbook(ws, wb, total_pages, max_rows, max_used_row, max_used_col, grid_params, output_path):
    from openpyxl.worksheet.pagebreak import Break
    from openpyxl.utils import get_column_letter

    ROW_PADDING = 1
    for pn in range(1, total_pages):
        ws.row_breaks.append(Break(id=pn * (max_rows + ROW_PADDING)))

    ws.page_setup.paperSize = grid_params.get('paper_size', 9)
    ws.page_setup.orientation = grid_params.get('orientation', 'portrait')
    ws.page_margins.left   = grid_params.get('margin_left',   0.43)
    ws.page_margins.right  = grid_params.get('margin_right',  0.43)
    ws.page_margins.top    = grid_params.get('margin_top',    0.41)
    ws.page_margins.bottom = grid_params.get('margin_bottom', 0.41)

    if max_used_row > 0 and max_used_col > 0:
        ws.print_area = f"A1:{get_column_letter(max_used_col)}{max_used_row}"

    wb.save(output_path)
    fix_empty_cell_type_attr(output_path)
    logger.info(f"[render_layout] Excel生成完了: {output_path} ({total_pages} ページ)")


def render_layout_to_xlsx(layout: list, grid_params: dict, output_path: str) -> None:
    COL_OFFSET = 1
    ROW_PADDING = 1
    grid_params = _refresh_render_params(grid_params)
    max_rows = grid_params['max_rows']
    default_font_size = grid_params.get('default_font_size', 7)
    font_name = grid_params.get('font_name', 'MS Gothic')

    wb, ws = _create_workbook(grid_params)
    max_used_row = 0
    max_used_col = 0

    for page_layout in layout:
        page_num = page_layout.get('page_number', 1)
        row_offset = (page_num - 1) * (max_rows + ROW_PADDING) + ROW_PADDING

        for elem in page_layout.get('elements', []):
            etype = elem.get('type')
            if etype == 'text':
                r, c = _place_text_element(ws, elem, row_offset, COL_OFFSET, default_font_size, font_name)
                max_used_row = max(max_used_row, r)
                max_used_col = max(max_used_col, c)
            elif etype == 'border_rect':
                er, ec = _place_border_element(ws, elem, row_offset, COL_OFFSET)
                # er/ec は排他的境界(最終セル+1)。print_area には実際に
                # 書き込まれた最終セル位置 (er-1, ec-1) を使う。
                max_used_row = max(max_used_row, er - 1)
                max_used_col = max(max_used_col, ec - 1)

    _finalize_workbook(ws, wb, len(layout), max_rows, max_used_row, max_used_col, grid_params, output_path)
