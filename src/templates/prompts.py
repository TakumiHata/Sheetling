"""
Sheetling 用の LLM プロンプト定義集。

  GEN_CODE_TEMPLATE        — _gen.py をスクリプトで直接生成するための string.Template
  VISUAL_REVIEW_PROMPT     — ビジョンLLMによる視覚的検証用プロンプト
  CODE_ERROR_FIXING_PROMPT — 生成コードのエラー修正プロンプト
"""
from string import Template

GRID_SIZES = {
    "small": {
        "col_width_mm": "3.48",
        "row_height_mm": "6.44",
        "max_cols": 54,
        "max_rows": 42,
        "excel_col_width": 1.71,
        "excel_row_height": 18.25,
        "margin_left": 0.43,
        "margin_right": 0.43,
        "margin_top": 0.41,
        "margin_bottom": 0.41,
        "default_font_size": 7,
        "font_name": "MS Gothic",
        "position_tolerance_cells": "1〜2",
    },
    "medium": {
        "col_width_mm": "6.0",
        "row_height_mm": "6.0",
        "max_cols": 36,
        "max_rows": 50,
        "excel_col_width": 2.53,
        "excel_row_height": 17.01,
        "margin_left": 0.47,
        "margin_right": 0.47,
        "margin_top": 0.41,
        "margin_bottom": 0.41,
        "default_font_size": 9,
        "position_tolerance_cells": "1",
    },
    "large": {
        "col_width_mm": "8.0",
        "row_height_mm": "8.0",
        "max_cols": 26,
        "max_rows": 38,
        "excel_col_width": 3.61,
        "excel_row_height": 22.68,
        "margin_left": 0.51,
        "margin_right": 0.51,
        "margin_top": 0.49,
        "margin_bottom": 0.49,
        "default_font_size": 11,
        "position_tolerance_cells": "1",
    },
    "pattern_1": {
        "col_width_mm": "3.48",
        "row_height_mm": "6.44",
        "max_cols": 54,
        "max_rows": 42,
        "excel_col_width": 1.71,
        "excel_row_height": 18.25,
        "margin_left": 0.43,
        "margin_right": 0.43,
        "margin_top": 0.41,
        "margin_bottom": 0.41,
        "default_font_size": 7,
        "position_tolerance_cells": "1〜2",
    },
    "pattern_2": {
        "col_width_mm": "5.53",
        "row_height_mm": "6.44",
        "max_cols": 34,
        "max_rows": 42,
        "excel_col_width": 2.71,
        "excel_row_height": 18.25,
        "margin_left": 0.43,
        "margin_right": 0.43,
        "margin_top": 0.41,
        "margin_bottom": 0.41,
        "default_font_size": 7,
        "position_tolerance_cells": "1〜2",
    },
    # Sheetling-pre "small"と同一セル密度（A4縦: 54列×42行）
    "1pt": {
        "col_width_mm": "3.48",
        "row_height_mm": "6.44",
        "max_cols": 57,            # 54+3: 右余白を埋めて A4 幅に合わせる
        "max_rows": 42,
        "excel_col_width": 1.625,  # (1*8+5)/8: デスクトップExcel(MDW=8)で列幅1.00表示
        "excel_row_height": 18.25,
        "margin_left": 0.43,
        "margin_right": 0.43,
        "margin_top": 0.41,
        "margin_bottom": 0.41,
        "default_font_size": 7,
        "font_name": "MS Gothic",
        "position_tolerance_cells": "1〜2",
    },
    # 粗めセル密度（A4縦: 34列×42行）
    "2pt": {
        "col_width_mm": "6.18",
        "row_height_mm": "6.44",
        "max_cols": 34,            # 34×2.71 MDU ≈ A4幅
        "max_rows": 42,
        "excel_col_width": 2.625,  # (2*8+5)/8: デスクトップExcel(MDW=8)で列幅2.00表示
        "excel_row_height": 18.25,
        "margin_left": 0.43,
        "margin_right": 0.43,
        "margin_top": 0.41,
        "margin_bottom": 0.41,
        "default_font_size": 6,
        "font_name": "MS Gothic",
        "position_tolerance_cells": "1",  # 4mm/セルと粗いため厳しく
    },
}


# ---------------------------------------------------------------------------
# 自動生成テンプレート（LLM不要）
# ---------------------------------------------------------------------------

# _gen.py をスクリプトで直接生成するための string.Template。
# $変数名 が grid_params と pdf_name で置換される。
# Python コード内の { } はそのまま使えるため .format() より安全。
GEN_CODE_TEMPLATE = Template("""\
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

MAX_ROWS = $max_rows
MAX_COLS = $max_cols
COL_OFFSET = 1
ROW_PADDING = 2
EXCEL_COL_WIDTH = $excel_col_width
EXCEL_ROW_HEIGHT = $excel_row_height
DEFAULT_FONT_SIZE = $default_font_size

_dir = Path(__file__).parent
data = json.loads((_dir / "${layout_json_name}").read_text(encoding="utf-8"))

wb = Workbook()
ws = wb.active
thin = Side(style='thin')
total_pages = len(data)

# シート全体のデフォルト列幅・行高さを設定（全セルに方眼サイズを適用）
ws.sheet_format.defaultColWidth = EXCEL_COL_WIDTH
ws.sheet_format.defaultRowHeight = EXCEL_ROW_HEIGHT
ws.sheet_format.customHeight = True


def apply_border(ws, s_row, e_row, s_col, e_col, borders):
    has_top    = borders.get("top",    True)
    has_bottom = borders.get("bottom", True)
    has_left   = borders.get("left",   True)
    has_right  = borders.get("right",  True)
    for r in range(s_row, e_row):
        for c in range(s_col, e_col):
            target = ws.cell(row=r, column=c)
            try:
                target.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            except AttributeError:
                pass
            top    = thin if (r == s_row and has_top)    else None
            bottom = thin if (r == e_row - 1 and has_bottom) else None
            left   = thin if (c == s_col and has_left)   else None
            right  = thin if (c == e_col - 1 and has_right)  else None
            try:
                target.border = Border(top=top, bottom=bottom, left=left, right=right)
            except AttributeError:
                pass


max_used_row = MAX_ROWS * total_pages + ROW_PADDING
max_used_col = MAX_COLS + COL_OFFSET

for page in data:
    page_number = page["page_number"]
    row_offset = (page_number - 1) * MAX_ROWS + ROW_PADDING

    if page_number > 1:
        ws.row_breaks.append(Break(id=(page_number - 1) * MAX_ROWS + ROW_PADDING))

    for item in page["elements"]:
        if item["type"] == "text":
            r = item["row"] + row_offset
            c = item["col"] + COL_OFFSET
            try:
                cell = ws.cell(row=r, column=c)
                cell.value = item["content"]
                if item.get("is_vertical"):
                    cell.alignment = Alignment(text_rotation=255, vertical='top', wrap_text=False)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                font_kwargs = {"size": item.get("font_size") or DEFAULT_FONT_SIZE}
                if item.get("font_color"):
                    font_kwargs["color"] = item["font_color"]
                if item.get("font_name"):
                    font_kwargs["name"] = item["font_name"]
                cell.font = Font(**font_kwargs)
                max_used_row = max(max_used_row, r)
                max_used_col = max(max_used_col, c)
            except AttributeError:
                pass

        elif item["type"] == "border_rect":
            s_row = item["row"] + row_offset
            e_row = item["end_row"] + row_offset
            s_col = item["col"] + COL_OFFSET
            e_col = item["end_col"] + COL_OFFSET
            apply_border(ws, s_row, e_row, s_col, e_col,
                         item.get("borders", {"top": True, "bottom": True, "left": True, "right": True}))
            max_used_row = max(max_used_row, e_row)
            max_used_col = max(max_used_col, e_col)

ws.page_setup.paperSize = $paper_size
ws.page_setup.orientation = '$orientation'
ws.page_margins.left = $margin_left
ws.page_margins.right = $margin_right
ws.page_margins.top = $margin_top
ws.page_margins.bottom = $margin_bottom
ws.print_area = f"A1:{get_column_letter(max_used_col)}{max_used_row}"
ws.page_setup.scale = 100

wb.save("output.xlsx")
""")


# ビジョンLLM（画像入力対応AIチャット）向けの視覚的検証プロンプト。
# JSONは渡さず、PDFページ画像 + 自動生成した罫線プレビュー画像の2枚比較で罫線差分を検出する。
# ページごとに生成し、対応する2枚のPNG画像と一緒にLLMに投入して使用する。
VISUAL_REVIEW_PROMPT = """\
以下の2枚の画像を比較し、**罫線（枠線）の明確な過不足のみ**を報告してください。

- 画像1: PDFの原本（ページ {page_number}）
- 画像2: 自動生成した罫線プレビュー（ページ {page_number}）

## 画像2（罫線プレビュー）の見方

- 薄いグレーの縦横線はグリッド背景線です。**罫線ではありません。無視してください。**
- 太い黒線のみが罫線（枠線）です。
- グリッドサイズ: {max_rows} 行 × {max_cols} 列（1マス = {col_width_mm}mm × {row_height_mm}mm）

### 座標の読み方（重要）

赤いラベルは **セルの中央** に表示されており、**ラベルの数値 = JSON の `row`/`col` 値** に直接対応します。

- ラベル `1` のセル → `col=1`（または `row=1`）
- ラベル `6` のセル → `col=6`（または `row=6`）
- ラベルのないセルはその前後の数値から数えてください（例: ラベル `6` の3つ右 → `col=9`）
- 左上セルが `(row=1, col=1)`、右・下方向に増加します

## 判定基準（厳格に守ってください）

**報告してよいもの（明確な差異のみ）**
- PDF に明らかに存在するが、プレビューに全く描画されていない罫線 → `add_border`
- プレビューに描画されているが、PDF には明らかに存在しない罫線 → `remove_border`

**報告してはいけないもの（無視してください）**
- テキスト・文字の差異（フォント・配置・内容の違いはすべて無視）
- グリッド背景線（薄いグレー線）
- 罫線の位置が {position_tolerance_cells} セルずれている程度の微小なズレ
- PDF の薄い罫線・飾り線・影など、Excel で表現不要な装飾的な線
- すでにプレビューに描画されている罫線を「位置修正」するような操作
- 判断に迷う・曖昧な差異（確信が持てない場合は報告しない）

**`remove_border` は特に慎重に使ってください。**
プレビューに罫線があり、PDFにも類似した線がある場合は、削除を提案しないでください。
明らかに余分な罫線（PDFのどこにも対応する線がない）にのみ使用してください。

## 出力形式

差異がない、または軽微な場合は `{{"corrections": []}}` のみ出力してください。

### フィールド名の厳守
- 列の終端: **`end_col`**（`col_end` は誤り）
- 行の終端: **`end_row`**（`row_end` は誤り）

### フォーマット

```json
{{
  "corrections": [
    {{
      "action": "add_border",
      "page": {page_number},
      "row": <開始行>,
      "end_row": <終了行>,
      "col": <開始列>,
      "end_col": <終了列>,
      "borders": {{"top": true, "bottom": true, "left": true, "right": true}}
    }},
    {{
      "action": "remove_border",
      "page": {page_number},
      "row": <開始行>,
      "end_row": <終了行>,
      "col": <開始列>,
      "end_col": <終了列>
    }}
  ]
}}
```

### 記入例（下線のみ追加する場合）
```json
{{"corrections": [{{"action": "add_border", "page": {page_number}, "row": 5, "end_row": 6, "col": 3, "end_col": 12, "borders": {{"bottom": true, "top": false, "left": false, "right": false}}}}]}}
```

【最重要】出力はJSONのみ。説明文・前置き・コードブロック記号（```）は一切不要です。
【最重要】疑わしい差異は報告しない。確実な差異のみ報告する。
【最重要】フィールド名は `end_col`・`end_row` を使用すること（`col_end`・`row_end` は誤り）。"""


# ---------------------------------------------------------------------------
CODE_ERROR_FIXING_PROMPT = """あなたはPythonプログラミングのエキスパートです。
以下のExcel生成用Pythonスクリプトを実行したところ、エラーが発生しました（または期待する結果が得られませんでした）。
エラーメッセージと現在のコードを分析し、問題を修正した新しいPythonスクリプトを生成してください。

【最重要：Python文法の遵守】
修正後のコード内では、必ず Python の予約語である `True`/`False` （先頭大文字）を使用してください。小文字の `true`/`false` を含めないでください。

【エラーメッセージ・実行結果】
{error_msg}

【現在のコード】
```python
{code}
```
【既知のエラーパターンと修正方法】
- `AttributeError: 'dict' object has no attribute 'to_tree'`
  → `ws.page_margins = {{...}}` のように dict を代入している。以下の属性代入形式に修正すること:
  ```python
  ws.page_margins.left = 0.47
  ws.page_margins.right = 0.47
  ws.page_margins.top = 0.41
  ws.page_margins.bottom = 0.41
  ws.page_margins.header = 0.0
  ws.page_margins.footer = 0.0
  ```
- `AttributeError: 'MergedCell' object attribute 'value' is read-only`
  → セルへの値代入や `merge_cells` 等の処理を `try...except AttributeError: pass` で囲むこと。

【最重要】出力は、修正済みの実行可能な Python コードのみをマークダウンのコードブロック (```python ... ```) で出力してください。前後の挨拶、解説、謝罪、思考プロセスなどは一切不要です。
【最重要】出力するコード内には、`[cite: ...]` のような参照タグやアノテーション（例: ` [cite: 271]`）を絶対に含めないでください。SyntaxErrorの原因となります。純粋で実行可能なPythonコードのみを出力してください。また、コードブロックの外にいかなるテキストも記述しないでください。"""

TABLE_ANCHOR_PROMPT = """あなたはPDF解析データからExcelレイアウト仕様を生成する設計者です。

## 座標について（最重要）

入力データの各要素には **事前計算済みのExcel座標** が付与されています。
独自に座標を計算せず、必ず以下のフィールドをそのまま参照してください。

| フィールド | 意味 |
|---|---|
| `words[i]._row` | テキストのExcel行番号 |
| `words[i]._col` | テキストのExcel列番号（開始） |
| `words[i].is_vertical` | 縦文字かどうか（`true` の場合のみ存在） |
| `words[i]._end_row` | 縦文字の下端行番号（`is_vertical=true` の場合のみ存在） |
| `rects[i]._row` / `_end_row` | 矩形の開始・終了行番号（テーブル外） |
| `rects[i]._col` / `_end_col` | 矩形の開始・終了列番号（テーブル外） |
| `rects[i]._borders` | 矩形枠の各辺に罫線があるか `{{top, bottom, left, right}}` |
| `table_border_rects[i]._row` / `_end_row` | テーブルセルの開始・終了行番号 |
| `table_border_rects[i]._col` / `_end_col` | テーブルセルの開始・終了列番号 |
| `table_border_rects[i]._borders` | セルの各辺に罫線があるか `{{top, bottom, left, right}}` |

## 複数ページ処理（最重要）

入力データには複数ページが含まれる場合があります。

- 出力JSONは `page_number` ごとに **必ず独立したオブジェクト** として出力すること
- ページ1の要素はページ1のオブジェクト内、ページ2の要素はページ2のオブジェクト内に格納する
- **あるページのコンテンツを別のページのオブジェクトに混入させることは絶対禁止**
- 各ページの座標（`_row`, `_col`）はそのページ内でリセットされた値（1始まり）である。ページをまたいで座標を統合・変換しないこと

## 処理手順

### Step 1: border_rect 要素の生成

**テーブルセル（`table_border_rects`）：**
- `table_border_rects` の各エントリを `border_rect` 要素として出力する
- 各エントリの `_row`, `_end_row`, `_col`, `_end_col` をそのまま使用する
- 各エントリの `_borders` を `borders` フィールドとしてそのまま出力する
- **隣接するセル同士で共有辺の `_borders` が両方 `false` の場合、それらは視覚的に1つのセルである可能性が高い。確信が持てる場合は統合して1つの `border_rect` にまとめよ**（統合後の row/col は最小値、end_row/end_col は最大値、borders は外周辺のみ true）

**テーブル外の矩形（`rects`）：**
- 各 rect の `_row`, `_col`, `_end_row`, `_end_col` をそのまま使用する
- 各 rect の `_borders` を `borders` フィールドとしてそのまま出力する

### Step 2: テキスト要素の生成

`words` を **`(_row, _col)` の組み合わせ** でグループ化し、同一 `(_row, _col)` のwordを1つの `text` 要素にまとめる。
これにより、テーブル内の各列セルが別々の `text` 要素として保持される（行単位の結合は禁止）。

- `row` = グループの `_row` 値
- `col` = グループの `_col` 値
- `content` = グループ内の `text` フィールドを左から順に**そのまま**結合したテキスト（日本語文字＝漢字・ひらがな・カタカナを含む場合はスペースなし、英数字のみの場合は半角スペースで結合）
  - **【厳守】`content` は入力 `words[i].text` の値をそのまま使用すること。要約・省略・言い換え・創作は絶対禁止。**
  - **【厳守】入力に存在しない文字を `content` に追加することは絶対禁止。**
- `end_col` = `col` + `content` の文字数（概算）
- `font_color` = グループ内最初のwordの `font_color`（存在し、かつ黒 `"000000"` でない場合のみ含める）
- `font_size` = グループ内最初のwordの `font_size`（存在する場合のみ含める）

**縦文字（`is_vertical=true`）の扱い：**
- `is_vertical=true` のwordは他のwordとグループ化せず、単独で `text` 要素にする
- `is_vertical: true` を要素に含める
- `_end_row` が存在する場合は `end_row` として含める（縦方向の占有範囲）
- `end_col` は `col + 1`（縦文字は1列幅）

## 出力フォーマット

複数ページの場合は以下のように **ページ数分のオブジェクトを配列に含める**。各ページの要素は必ず対応する `page_number` のオブジェクト内にのみ格納すること。

[
  {{
    "page_number": 1,
    "elements": [
      {{
        "type": "text",
        "content": "請求書",
        "row": 2,
        "col": 20,
        "end_col": 28,
        "font_color": "FF0000",
        "font_size": 14
      }},
      {{
        "type": "border_rect",
        "row": 8,
        "end_row": 9,
        "col": 18,
        "end_col": 27,
        "borders": {{"top": true, "bottom": true, "left": false, "right": true}}
      }}
    ]
  }},
  {{
    "page_number": 2,
    "elements": [
      {{
        "type": "text",
        "content": "明細",
        "row": 3,
        "col": 5,
        "end_col": 7
      }},
      {{
        "type": "border_rect",
        "row": 5,
        "end_row": 10,
        "col": 3,
        "end_col": 25,
        "borders": {{"top": true, "bottom": true, "left": true, "right": true}}
      }}
    ]
  }}
]

入力データ:
{input_data}

【最重要】出力は `[` で始まり `]` で終わる純粋なJSON配列文字列のみとしてください。Markdownのコードブロック(```json等)、前ворю後の説明文、思考プロセス、検証コメントは一切含めないでください。JSON以外の文字を1文字でも出力するとSTEP 1.5で処理できなくなります。"""


LAYOUT_REVIEW_PROMPT = """あなたはExcelレイアウト設計の検証者です。
STEP 1が生成したレイアウト仕様JSONを検証・補正し、同じフォーマットの修正済みJSONを出力してください。

## 検証・修正項目

### 1. border_rect の座標整合性
- `row > end_row` または `col > end_col` となっている場合は値を入れ替えて修正する
- `row == end_row` かつ `col == end_col`（1×1セル）のborder_rectは削除する

### 2. 重複するborder_rectの除去
- `row`, `end_row`, `col`, `end_col` がすべて同一のborder_rectが複数ある場合、1つだけ残す

### 2.5. border_rectのbordersフィールド
- `borders` フィールドが存在する場合、値を変更せずそのまま保持すること
- STEP 1 で隣接セルが統合された場合、統合後の `borders` は外周辺（視覚的に線がある辺）のみ true にすること

### 2.7. ページ間要素混入の検出・修正（最重要）
- 元のPDF解析データの `page_number` と照合し、あるページのコンテンツが別のページの `elements` に混入していないか確認する
- 例：元データのページ2に存在する `words` がSTEP 1出力のページ1の `elements` に含まれている場合、それをページ1から削除しページ2の `elements` に移動する
- 判定基準：元データの `pages[N].words[i].text` が STEP 1 出力の `page_number: M`（M ≠ N+1）の elements に含まれている場合は混入

### 3. テキスト内容の正確性検証（最重要）
- 各 `text` 要素の `content` が、元データの対応する `words[i].text` の値を正確に使用しているか確認する
- `content` に入力 `words` に存在しない文字列が含まれている場合（要約・省略・言い換え・創作）、元の `words[i].text` を使った正確な値に修正する
- `content` が元データより短く省略されている場合も同様に修正する

### 4. テキスト要素の欠落補完
以下の元PDF解析データの `words` と照合し、レイアウトJSONに欠落しているテキストを補完する：
- 元データの各ページの `words` を確認し、**対応するページ**の `text` 要素に含まれていない `text` フィールドの内容を探す（ページ番号を必ず照合すること）
- 欠落が確認された場合、該当wordの `_row` を `row`、`_col` を `col` として **同じページ番号** の `text` 要素を追加する
- `content` は該当wordの `text` フィールドの値をそのまま使用する（要約・省略禁止）
- `end_col` = `col` + `text` の文字数（概算）
- 空白・記号のみのword（`text` が空文字、スペースのみ、または1文字以下の記号）は無視してよい

### 5. text 要素の重複除去
- 同一ページ内で `row` と `col` が同じ `text` 要素が複数ある場合、`content` を結合して1つにまとめる（日本語文字を含む場合はスペースなし、英数字のみの場合は半角スペース）

### 6. 座標のクランプ
- `row`, `end_row` が {max_rows} を超えている場合は {max_rows} に切り詰める
- `col`, `end_col` が {max_cols} を超えている場合は {max_cols} に切り詰める

## 元のPDF解析データ（参照用）

{input_data}

## STEP 1の出力（検証・修正対象）

{step1_output}

【最重要】出力は `[` で始まり `]` で終わる純粋なJSON配列文字列のみとしてください。Markdownのコードブロック(```json等)、前後の説明文、思考プロセス、検証コメントは一切含めないでください。JSON以外の文字を1文字でも出力するとgenerate処理できなくなります。"""
