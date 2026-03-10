"""
AI出力のPythonソースコードを実行し、3シートのExcelファイルを生成するモジュール。

- 1シート目: AI生成コードによるExcel描画（PDFの再現）
- 2シート目: カラーコード・フォント情報の一覧
- 3シート目: PDF画像の添付
"""

import os
import traceback

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

from src.core.config import config
from src.utils.logger import get_logger

logger = get_logger(__name__)


class Executor:
    """AI出力のPythonコードを実行し、2シート構成のExcelを生成する"""

    def __init__(self):
        self.col_width = config.excel.col_width_chars
        self.row_height = config.excel.row_height_pt
        self.max_cols = config.grid.target_cols
        self.max_rows = config.grid.target_rows

    def execute(
        self,
        gen_py_path: str,
        output_xlsx_path: str,
        num_pages: int = 1,
        page_heights: list[float] = None,
        page_breaks: list[int] = None,
    ) -> str:
        """
        AI生成のPythonソースを実行し、Excelを出力する。

        Args:
            gen_py_path: AI出力の .py ファイルパス
            output_xlsx_path: 出力する .xlsx ファイルパス
            num_pages: PDFの総ページ数
            page_heights: 各ページの実際高さリスト(pt)。改ページ行番号計算に使用。

        Returns:
            出力されたExcelファイルパス
        """
        logger.info(f"--- Executing AI-generated code: {gen_py_path} ---")

        # page_heightsから総行数を計算
        import math as _math
        if page_heights:
            total_rows = _math.ceil(sum(page_heights) / self.row_height)
        else:
            total_rows = self.max_rows * max(num_pages, 1)

        # Workbook初期化
        wb = self._create_workbook(total_rows=total_rows)
        ws = wb.active

        # AI生成コードの実行 (total_rowsを渡して不足分を補えるようにする)
        self._execute_generated_code(gen_py_path, wb, ws, page_heights=page_heights, total_rows=total_rows, page_breaks=page_breaks)

        # 保存
        os.makedirs(os.path.dirname(os.path.abspath(output_xlsx_path)), exist_ok=True)
        wb.save(output_xlsx_path)
        logger.info(f"✅ Excel saved: {output_xlsx_path}")

        return output_xlsx_path

    def _create_workbook(self, total_rows: int = None) -> openpyxl.Workbook:
        """実際の総行数に合わせた方眼設定済みWorkbookを作成する"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "変換結果"

        # 列幅（文字幅基準）の変動を防ぐため、標準フォントを Arial 11pt に固定する
        for s in wb._named_styles:
            if s.name == "Normal":
                s.font = Font(name="Arial", size=11)
                break

        # ページ設定（A4）
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

        # 印刷余白（方眼紙を最大限活かすため最小限に設定）
        ws.page_margins.left = 0.1
        ws.page_margins.right = 0.1
        ws.page_margins.top = 0.3
        ws.page_margins.bottom = 0.3
        ws.page_margins.header = 0.2
        ws.page_margins.footer = 0.2

        # ページフィット設定
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        # 1ページのみの場合は高さも1ページに収める。複数ページの場合は縦に伸ばす。
        ws.page_setup.fitToHeight = 1 if total_rows <= self.max_rows else 0

        # 方眼の列幅を設定
        for col_idx in range(1, self.max_cols + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = self.col_width

        # 実際の総行数分の行高さを設定（2ページ目以降が未設定だと方眼比率が崩れる）
        rows_to_set = total_rows if total_rows else self.max_rows
        for row_idx in range(1, rows_to_set + 1):
            ws.row_dimensions[row_idx].height = self.row_height

        return wb

    def _execute_generated_code(self, gen_py_path: str, wb, ws, page_heights: list[float] = None, total_rows: int = None, page_breaks: list[int] = None):
        """AI生成のPythonコードを読み込んで実行する"""
        import math as _math
        logger.info(f"Loading generated code: {gen_py_path}")

        # AIが生成したコードを文字列として読み込む
        with open(gen_py_path, "r", encoding="utf-8") as f:
            code = f.read()

        # 動的実行のための独立した名前空間を用意する
        exec_globals = {
            "__builtins__": __builtins__,
        }

        try:
            # コード文字列をコンパイルし、用意した名前空間上で実行する
            exec(compile(code, gen_py_path, "exec"), exec_globals)

            if "generate" not in exec_globals:
                raise RuntimeError(
                    f"AI生成コードに `generate(wb, ws)` 関数が定義されていません: {gen_py_path}"
                )

            # generate関数を呼び出し
            exec_globals["generate"](wb, ws)
            logger.info("✅ AI generated code executed successfully (Sheet 1)")

            # --- 改ページを挿入 ---
            if page_breaks:
                for br in page_breaks:
                    if br > 0:
                        ws.row_breaks.append(Break(id=br))
                        logger.info(f"  改ページ at row {br} (from extracted page_breaks)")
            elif page_heights and len(page_heights) > 1:
                # 従来通りのポイント計算（フォールバック）
                cumulative_pt = 0.0
                for ph in page_heights[:-1]:  # 最終ページの後は改ページ不要
                    cumulative_pt += ph
                    break_row = _math.ceil(cumulative_pt / self.row_height)
                    ws.row_breaks.append(Break(id=break_row))
                    logger.info(f"  改ページ at row {break_row} (cumulative {cumulative_pt:.1f}pt)")
            else:
                # page_heightsがない場合は従来の固定値フォールバック
                max_row = ws.max_row
                if max_row > self.max_rows:
                    for r in range(self.max_rows, max_row, self.max_rows):
                        ws.row_breaks.append(Break(id=r))

            # 実際の書き込み済みの最大行まで方眼高さを保証する
            max_r = ws.max_row
            if total_rows and max_r > total_rows:
                for row_idx in range(total_rows + 1, max_r + 1):
                    ws.row_dimensions[row_idx].height = self.row_height

            # --- 印刷範囲（Print Area）をデータが存在する最大行から設定 ---
            last_col_letter = get_column_letter(self.max_cols)
            ws.print_area = f"A1:{last_col_letter}{max_r}"
            logger.info(f"✅ Set exact print area to: {ws.print_area}")


        except Exception as e:
            logger.error(f"❌ AI generated code execution failed: {e}")
            logger.error(traceback.format_exc())

            # エラー発生時に内容がわかるよう、最初の数行分だけセルの高さを広げる
            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 20
            ws.row_dimensions[3].height = 60

            ws["A1"] = "⚠ AI生成コードの実行に失敗しました"
            ws["A1"].font = Font(color="FF0000", size=14, bold=True)

            ws["A2"] = f"エラー種別: {type(e).__name__}"
            ws["A2"].font = Font(size=10)

            ws["A3"] = f"詳細: {str(e)}"
            ws["A3"].font = Font(size=10)
            ws["A3"].alignment = Alignment(wrap_text=True)

            # エラー内容が見えるようにA列のみ幅を拡大
            ws.column_dimensions["A"].width = 80

