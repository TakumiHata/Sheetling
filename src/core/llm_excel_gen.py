import json
import os
import pandas as pd
from pathlib import Path
from src.utils.logger import get_logger

logger = get_logger(__name__)

class LLMExcelGenerator:
    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(self, master_json_path: str, dl_md_path: str, mid_md_path: str):
        """
        ハイブリッド解析結果（JSON/MD）を基に、LLM向けのプロンプトおよびExcel設計図を生成し、
        最終的なExcelバイナリを出力する。
        """
        base_name = Path(master_json_path).stem.replace("_master", "")
        logger.info(f"Generating Excel design and prompt from {master_json_path}")
        
        # 1. マスターJSONの読み込み
        with open(master_json_path, "r", encoding="utf-8") as f:
            master_data = json.load(f)
            
        # 2. LLM向けのプロンプト（指示書）の構築
        prompt = self._build_prompt(master_data, dl_md_path, mid_md_path)
        
        # プロンプトをファイルとして保存（デバッグ・LLM入力用）
        prompt_path = self.output_dir / f"{base_name}_prompt.txt"
        with open(prompt_path, "w", encoding="utf-8") as f:
            f.write(prompt)
            
        # 3. Excelバイナリの生成 (openpyxlによる方眼レイアウトの実装)
        excel_path = self._assemble_excel(master_data, base_name)
        
        logger.info(f"Prompt context and Excel generated at: {excel_path}")
        return str(excel_path)

    def _build_prompt(self, master_data: dict, dl_md_path: str, mid_md_path: str):
        """
        LLMに「方眼Excelの設計図」を書かせるための高度なプロンプトを構築する。
        """
        prompt = [
            "# 指示",
            "あなたはPDF解析結果から「方眼Excel（5pt単位）」を再構築するエキスパートです。",
            "提供されたマスターJSON（レイアウト構造）とMarkdown（テキスト内容）を基に、以下の3ステップでExcel生成命令セット（JSON形式）を作成してください。",
            "",
            "## 生成ステップ",
            "1. **レイアウトの構築**: マスターJSONの `type: box` を基に、セルの背景色（HEX）と範囲を特定します。",
            "2. **テキストの流し込み**: マスターJSONの `type: text` と **高精度テキストMD** の内容を完全に同期させ、文字漏れ・誤字がないように流し込んでください。特に表内のデータや金額は1文字も漏らさず記載してください。",
            "3. **最適化**: 連続するセルは積極的に「結合（Merge）」し、フォントサイズや折り返し設定を適切に指定してください。",
            "",
            "## ルール",
            "1. **A4サイズ準拠**: 出力はA4サイズ1ページ（または複数ページ）に収まるようにレイアウトを調整してください。",
            "2. **余白の最適化**: 印刷時に内容が切れないよう、上下左右に適切な余白（Margin）を考慮した命令を作成してください。",
            "3. **グリッド指定**: 座標は5pt単位のグリッド（Grid Row/Col）で厳密に指定してください。",
            "4. **背景色**: 必ずマスターJSONの色彩情報を優先してください。",
            "5. **結合の優先**: テキストが複数のセルに跨る場合や、見た目の整合性を保つ場合は積極的に `action: merge` を使用してください。",
            "",
            "## 解析データ概要",
            f"- PDF名: {master_data['pdf_name']}",
            f"- 方眼単位: {master_data['grid_size']} pt",
            f"- ページ数: {len(master_data['pages'])}",
            "",
            "## 参考リソース",
            f"- 構造MD: {dl_md_path}",
            f"- 高精度テキストMD: {mid_md_path}",
            "",
            "## 命令セットの出力形式（JSONのみを出力）",
            "```json",
            "[",
            "  { \"action\": \"merge\", \"range\": \"B2:E4\" },",
            "  { \"action\": \"fill\", \"range\": \"B2:E4\", \"color\": \"#F2F2F2\" },",
            "  { \"action\": \"write\", \"range\": \"B2\", \"value\": \"項目名\", \"align\": \"center\" }",
            "]",
            "```"
        ]
        return "\n".join(prompt)

    def _assemble_excel(self, master_data: dict, base_name: str):
        """
        openpyxlを使用して、解析されたグリッド情報をベースにExcelファイルを構築する。
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Page 1"
        
        # 方眼のセットアップ（全列・全行の幅・高さを固定）
        # 列幅: 2.14 unit ≒ 5pt (方眼紙の最小単位)
        for i in range(1, 200):
            ws.column_dimensions[get_column_letter(i)].width = 2.0
        for i in range(1, 400):
            ws.row_dimensions[i].height = 12.0
            
        # 最初のページのエレメントを流し込む
        page = master_data["pages"][0]
        # boxes(背景)を先に処理して、必要なら結合
        for i, elem in enumerate(page["elements"]):
            if i > 1000: break
            if elem["type"] == "box":
                gr = elem["grid_range"]
                c1, r1 = gr["start_col"] + 1, gr["start_row"] + 1
                c2, r2 = gr["end_col"] + 1, gr["end_row"] + 1
                c_start, c_end = min(c1, c2), max(c1, c2)
                r_start, r_end = min(r1, r2), max(r1, r2)
                
                # 背景色の反映
                color_hex = self._rgb_to_hex(elem.get("color"))
                if color_hex:
                    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    for r in range(r_start, r_end + 1):
                        for c in range(c_start, c_end + 1):
                            target_cell = ws.cell(row=r, column=c)
                            try:
                                target_cell.fill = fill
                            except AttributeError:
                                pass
                
                # 大きなボックス（容器）は結合を試みる（テキストとの競合は後でハンドリング）
                # 面積が一定以上（例: 4マス以上）なら結合
                if (c_end - c_start + 1) * (r_end - r_start + 1) >= 4:
                    try:
                        ws.merge_cells(start_row=r_start, start_column=c_start, end_row=r_end, end_column=c_end)
                    except Exception:
                        pass

        # 次にテキストを処理（既に結合されているエリアに書く場合は左上に書く）
        for i, elem in enumerate(page["elements"]):
            if i > 1000: break
            if elem["type"] == "text":
                gr = elem["grid_range"]
                c1, r1 = gr["start_col"] + 1, gr["start_row"] + 1
                c2, r2 = gr["end_col"] + 1, gr["end_row"] + 1
                c_start, c_end = min(c1, c2), max(c1, c2)
                r_start, r_end = min(r1, r2), max(r1, r2)

                # 値を書き込む（マージ済みの場合は左上にしか書けない）
                cell = ws.cell(row=r_start, column=c_start)
                try:
                    cell.value = elem["content"]
                except AttributeError:
                    pass
                
                # アライメント設定
                from openpyxl.styles import Alignment
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

                # 指定されたグリッド範囲を結合（まだ結合されていない場合のみ）
                if c_start != c_end or r_start != r_end:
                    try:
                        ws.merge_cells(start_row=r_start, start_column=c_start, end_row=r_end, end_column=c_end)
                    except Exception:
                        pass

        excel_path = self.output_dir / f"{base_name}.xlsx"
        
        # A4サイズ・印刷設定の反映
        # https://openpyxl.readthedocs.io/en/stable/print_settings.html
        from openpyxl.worksheet.page import PageMargins
        
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        
        # 余白調整 (0.5 inch ≒ 1.27cm)
        ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5, header=0.3, footer=0.3)
        
        # A4の幅に収めるためのスケーリング（自動調整）
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0 # 高さはページを跨いでも良い
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        # 印刷範囲の指定 (grid_cols, grid_rowsに基づく)
        max_col = page["grid_cols"]
        max_row = page["grid_rows"]
        ws.print_area = f"A1:{get_column_letter(max_col)}{max_row}"

        wb.save(excel_path)
        return str(excel_path)

    def _rgb_to_hex(self, rgb):
        """
        [R, G, B] 配列を HEX 形式に変換。
        """
        if not rgb or not isinstance(rgb, (list, tuple)) or len(rgb) != 3:
            return None
            
        # pdfplumberのrgbは通常0.0-1.0または0-255
        if all(isinstance(x, float) and x <= 1.0 for x in rgb):
            rgb = [int(x * 255) for x in rgb]
        return "{:02X}{:02X}{:02X}".format(rgb[0], rgb[1], rgb[2])
