"""
Sheetling パイプライン。

【全自動モード】
  auto: PDF解析 → レイアウトJSON自動生成 → _gen.py 生成 → Excel描画
  correct: ビジョンLLM修正指示を適用して Excel を再生成

【LLM協業モード（高精度）】
  extract: PDF解析 → STEP1/STEP1.5 プロンプト生成 → PDFページ画像出力
  fill:    STEP1.5 LLM出力 → テキスト補完 → layout.json 保存
  review:  layout.json → 罫線プレビュー画像 + VISUAL_REVIEW_PROMPT 生成
  generate: layout.json → Excel 直接描画（corrections 適用後）
"""

import json
import math
import re
from collections import defaultdict
from pathlib import Path

from src.parser.pdf_extractor import extract_pdf_data
from src.templates.prompts import (
    VISUAL_REVIEW_PROMPT, GRID_SIZES,
)
from src.utils.logger import get_logger


# MS Access / Windows 環境でよく使われるフォント名の正規化テーブル。
# PDF 埋め込み名はハイフン・スペース・大小文字が揺れるため、
# Excel が認識できる表記に統一する。
_FONT_ALIASES: dict = {
    'MS Gothic':     'MS Gothic',
    'MSGothic':      'MS Gothic',
    'MS PGothic':    'MS PGothic',
    'MSPGothic':     'MS PGothic',
    'MS Mincho':     'MS Mincho',
    'MSMincho':      'MS Mincho',
    'MS PMincho':    'MS PMincho',
    'MSPMincho':     'MS PMincho',
    'MS UI Gothic':  'MS UI Gothic',
    'MSUIGothic':    'MS UI Gothic',
    'Meiryo':        'Meiryo',
    'Meiryo UI':     'Meiryo UI',
    'MeiryoUI':      'Meiryo UI',
    'Yu Gothic':     'Yu Gothic',
    'YuGothic':      'Yu Gothic',
    'Yu Mincho':     'Yu Mincho',
    'YuMincho':      'Yu Mincho',
}


def _normalize_font_name(raw_name):
    """PDF フォント名を Excel に渡せる形式に整形する。
    1. bytes / bytes repr 文字列（pdfminer が返す非ASCII名）を除外
    2. サブセットプレフィックスを除去 (例: "ABCDEF+MS-Gothic" → "MS-Gothic")
    3. ハイフン区切りをスペースに変換 (例: "MS-Gothic" → "MS Gothic")
    4. _FONT_ALIASES でエイリアス解決（揺れ表記を正規表記に統一）
    Excel が認識できないフォント名はデフォルトフォントにフォールバックされる。"""
    if not raw_name:
        return None
    # bytes オブジェクトは使用不可
    if isinstance(raw_name, bytes):
        return None
    # pdfminer が bytes を str() した "b'...'" 形式の文字列も使用不可
    if isinstance(raw_name, str) and raw_name.startswith("b'"):
        return None
    # サブセットプレフィックスを除去 (例: "ABCDEF+MS-Gothic" → "MS-Gothic")
    name = re.sub(r'^[A-Z]{6}\+', '', raw_name)
    # ハイフン区切りをスペースに変換 (例: "MS-Gothic" → "MS Gothic")
    name = name.replace('-', ' ').strip()
    # エイリアステーブルで正規化（大文字小文字も考慮して前方一致）
    return _FONT_ALIASES.get(name, name) or None



def _to_monotone(idxs: list, max_val: int) -> list:
    """重複した整数インデックスを +1 でずらして単調増加にする。"""
    result = []
    prev = 0
    for idx in idxs:
        idx = max(idx, prev + 1)
        result.append(min(idx, max_val))
        prev = result[-1]
    return result


def _compute_grid_coords(page: dict, max_rows: int, max_cols: int) -> None:
    """
    PDF座標をExcel行・列番号に直接変換し、各要素にインプレースで付与する。
    A4固定・グリッド固定の前提で floor 除算を使うため、クラスタリング不要。
    grid_h ≈ 20pt の幅の中でのサブピクセルノイズ（±数pt）は自然に同一セルに収まる。
    """
    grid_h = page['height'] / max_rows
    grid_w = page['width']  / max_cols

    def to_row(y: float) -> int:
        return max(1, min(max_rows, 1 + int(float(y) / grid_h)))

    def to_col(x: float) -> int:
        return max(1, min(max_cols, 1 + int(float(x) / grid_w)))

    # words に行・列番号を付与
    # top < 0 または top > height はページ境界外のアーティファクトのためスキップ
    page_h = float(page['height'])
    for w in page['words']:
        t = float(w.get('top', 0))
        if t < 0 or t > page_h:
            continue
        w['_row'] = to_row(t)
        w['_col'] = to_col(w['x0'])
        if w.get('is_vertical') and 'bottom' in w:
            w['_end_row'] = to_row(w['bottom'])

    # 薄い矩形（線の太さ分）を統合して1つの矩形にする。
    # 4本の線（上辺H + 下辺H + 左辺V + 右辺V）で構成される枠を検出し、
    # 1つの完全な矩形に統合する。
    _LINE_THICKNESS = 3.0  # pt: これ未満の幅/高さは線とみなす
    _MERGE_TOL = 5.0  # pt: 端点がこの距離以内なら接続とみなす

    h_lines = []  # (idx, x0, x1, y)
    v_lines = []  # (idx, x, y0, y1)
    normal_rects = []
    for idx, r in enumerate(page['rects']):
        w = abs(r['x1'] - r['x0'])
        h = abs(r['bottom'] - r['top'])
        if w < _LINE_THICKNESS and h >= _LINE_THICKNESS:
            v_lines.append((idx, (r['x0'] + r['x1']) / 2, r['top'], r['bottom']))
        elif h < _LINE_THICKNESS and w >= _LINE_THICKNESS:
            h_lines.append((idx, r['x0'], r['x1'], (r['top'] + r['bottom']) / 2))
        else:
            normal_rects.append(idx)

    # 線を統合して矩形を構築
    used_indices: set = set()
    merged_rects: list = []
    for hi_top in range(len(h_lines)):
        if h_lines[hi_top][0] in used_indices:
            continue
        _, hx0_t, hx1_t, hy_t = h_lines[hi_top]
        # 同じ x 範囲で下にある横線を探す
        for hi_bot in range(len(h_lines)):
            if hi_bot == hi_top or h_lines[hi_bot][0] in used_indices:
                continue
            _, hx0_b, hx1_b, hy_b = h_lines[hi_bot]
            if hy_b <= hy_t:
                continue  # 上辺より上にある
            if abs(hx0_t - hx0_b) > _MERGE_TOL or abs(hx1_t - hx1_b) > _MERGE_TOL:
                continue  # x 範囲が合わない
            # 左辺の縦線を探す
            vl_found = None
            for vi in range(len(v_lines)):
                if v_lines[vi][0] in used_indices:
                    continue
                _, vx, vy0, vy1 = v_lines[vi]
                if (abs(vx - min(hx0_t, hx0_b)) < _MERGE_TOL and
                        abs(vy0 - hy_t) < _MERGE_TOL and abs(vy1 - hy_b) < _MERGE_TOL):
                    vl_found = vi
                    break
            # 右辺の縦線を探す
            vr_found = None
            for vi in range(len(v_lines)):
                if v_lines[vi][0] in used_indices:
                    continue
                _, vx, vy0, vy1 = v_lines[vi]
                if (abs(vx - max(hx1_t, hx1_b)) < _MERGE_TOL and
                        abs(vy0 - hy_t) < _MERGE_TOL and abs(vy1 - hy_b) < _MERGE_TOL):
                    vr_found = vi
                    break
            # 少なくとも2辺以上見つかれば矩形として統合
            if vl_found is not None or vr_found is not None:
                used_indices.add(h_lines[hi_top][0])
                used_indices.add(h_lines[hi_bot][0])
                if vl_found is not None:
                    used_indices.add(v_lines[vl_found][0])
                if vr_found is not None:
                    used_indices.add(v_lines[vr_found][0])
                x0 = min(hx0_t, hx0_b)
                x1 = max(hx1_t, hx1_b)
                if vl_found is not None:
                    x0 = min(x0, v_lines[vl_found][1])
                if vr_found is not None:
                    x1 = max(x1, v_lines[vr_found][1])
                merged_rects.append({
                    'x0': x0, 'top': hy_t, 'x1': x1, 'bottom': hy_b,
                })
                break  # 上辺に対して1つの矩形を統合したら次へ

    # 統合されなかった線は個別に残す
    remaining = [page['rects'][i] for i in range(len(page['rects']))
                 if i not in used_indices]
    # 統合された矩形を追加
    remaining.extend(merged_rects)
    page['rects'] = remaining

    # rects に行・列番号を付与
    # end_row/end_col は +1 してテキストが枠内に収まるようにする。
    # rects はテーブル外の独立した矩形のため、隣接セルとの重なりは問題にならない。
    for r in page['rects']:
        r['_row']     = to_row(r['top'])
        r['_end_row'] = to_row(r['bottom']) + 1
        r['_col']     = to_col(r['x0'])
        r['_end_col'] = to_col(r['x1']) + 1

    # テーブル内に含まれる rects を除外（table_border_rects で代替するため）
    tol = 3.0
    table_bboxes = page.get('table_bboxes', [])

    def is_inside_table(r: dict) -> bool:
        for bbox in table_bboxes:
            if (r['x0'] >= bbox[0] - tol and r['x1'] <= bbox[2] + tol and
                    r['top'] >= bbox[1] - tol and r['bottom'] <= bbox[3] + tol):
                return True
        return False

    page['rects'] = [r for r in page['rects'] if not is_inside_table(r)]

    # テーブルの cells_2d から border_rect を生成。
    # bbox がある（None でない）セルのみ罫線を描画する。
    # bbox が None のセルは結合延長であり罫線を引かない。
    # これにより結合セル内部の不要な縦線・横線が除去される。
    table_border_rects = []
    for _cells_2d in page.get('table_cells', []):
        if not _cells_2d:
            continue
        for ri, row_cells in enumerate(_cells_2d):
            for ci, cb in enumerate(row_cells):
                if cb is None:
                    continue  # 結合延長 → 罫線なし
                r  = max(1, to_row(float(cb['top'])))
                er = max(r + 1, min(max_rows, to_row(float(cb['bottom']))))
                c  = max(1, to_col(float(cb['x0'])))
                ec = max(c + 1, min(max_cols, to_col(float(cb['x1']))))
                table_border_rects.append({
                    '_row': r, '_end_row': er,
                    '_col': c, '_end_col': ec,
                    '_pdf_x0': float(cb['x0']), '_pdf_top': float(cb['top']),
                    '_pdf_x1': float(cb['x1']), '_pdf_bottom': float(cb['bottom']),
                    '_borders': {'top': True, 'bottom': True, 'left': True, 'right': True},
                })
    page['table_border_rects'] = table_border_rects

    # rects に _borders を付与（4辺全て True — テーブル外の矩形・罫線）
    for rect in page['rects']:
        rect['_borders'] = {'top': True, 'bottom': True, 'left': True, 'right': True}

    # ---- PDF 余白分の空き行を除去するため _row を正規化 --------------------------------
    all_rows = (
        [w['_row'] for w in page['words'] if '_row' in w]
        + [r['_row'] for r in page['rects'] if '_row' in r]
        + [tbr['_row'] for tbr in page['table_border_rects']]
    )
    if all_rows:
        row_shift = min(all_rows) - 1
        if row_shift > 0:
            for w in page['words']:
                if '_row' in w:
                    w['_row'] -= row_shift
                    if '_end_row' in w:
                        w['_end_row'] -= row_shift
            for r in page['rects']:
                if '_row' in r:
                    r['_row'] -= row_shift
                    r['_end_row'] -= row_shift
            for tbr in page['table_border_rects']:
                tbr['_row'] -= row_shift
                tbr['_end_row'] -= row_shift

    # ---- LLM には渡さない（auto モードでは _auto_generate_layout が使用するため残す）------
    # table_cells / table_data / table_row_y_positions は _table_text_elements_from_2d で使うため
    # auto_layout 側で layout 生成後に削除する
    page.pop('h_edges', None)
    page.pop('v_edges', None)


# 旧コードにあった後処理（クラスタリング時代の補正。直接除算後は不要なため削除済み）:
#   - build_cluster_map + anchor_vals によるクラスタリング
#   - 同一視覚行ワードの top 正規化（_same_row_groups）
#   - テーブル列/行が同一グリッドに潰れた場合の後処理ループ
#   - テーブル底辺直下の空行挿入ロジック
# 直接除算では grid_h ≈ 20pt の幅の中でのノイズが自然に同一セルに収まるため不要。

logger = get_logger(__name__)


# ===========================================================================
# pre版から移植: LLM協業モード用ユーティリティ関数
# ===========================================================================

def _merge_table_border_rects(tbrs: list) -> list:
    """
    隣接セル間に境界線がない table_border_rects を統合し、結合セルを 1 つの
    border_rect として表現する。隣接整合性パスの後に呼ぶこと（共有辺の値が一致済み）。
    """
    # --- 縦方向マージ ---
    col_groups: dict = defaultdict(list)
    for tbr in tbrs:
        col_groups[(tbr['_col'], tbr['_end_col'])].append(tbr)

    v_merged: list = []
    for cells in col_groups.values():
        cells = sorted(cells, key=lambda c: c['_row'])
        stack = [dict(cells[0])]
        for cell in cells[1:]:
            prev = stack[-1]
            if prev['_end_row'] == cell['_row'] and not prev['_borders']['bottom']:
                prev['_end_row']           = cell['_end_row']
                prev['_pdf_bottom']        = cell['_pdf_bottom']
                prev['_borders']['bottom'] = cell['_borders']['bottom']
                prev['_borders']['left']   = prev['_borders']['left']  or cell['_borders']['left']
                prev['_borders']['right']  = prev['_borders']['right'] or cell['_borders']['right']
                prev['_outer_right']       = prev.get('_outer_right', False) or cell.get('_outer_right', False)
                prev['_major_right']       = prev.get('_major_right', False) or cell.get('_major_right', False)
            else:
                stack.append(dict(cell))
        v_merged.extend(stack)

    # --- 横方向マージ ---
    row_groups: dict = defaultdict(list)
    for tbr in v_merged:
        row_groups[(tbr['_row'], tbr['_end_row'])].append(tbr)

    h_merged: list = []
    for cells in row_groups.values():
        cells = sorted(cells, key=lambda c: c['_col'])
        stack = [dict(cells[0])]
        for cell in cells[1:]:
            prev = stack[-1]
            if prev['_end_col'] == cell['_col'] and not prev['_borders']['right']:
                prev['_end_col']           = cell['_end_col']
                prev['_pdf_x1']            = cell['_pdf_x1']
                prev['_borders']['right']  = cell['_borders']['right']
                prev['_borders']['top']    = prev['_borders']['top']    or cell['_borders']['top']
                prev['_borders']['bottom'] = prev['_borders']['bottom'] or cell['_borders']['bottom']
                prev['_outer_right']       = cell.get('_outer_right', False)
                prev['_major_right']       = cell.get('_major_right', False)
            else:
                stack.append(dict(cell))
        h_merged.extend(stack)

    return h_merged


def _fix_empty_cell_type_attr(xlsx_path: str) -> None:
    """
    openpyxl 3.1.x は値なしでスタイル（罫線）のみ設定されたセルに t="n" 属性を付与する。
    Excel Online はこれを不正な属性として修復処理（ブックが修復されました）を行い、
    罫線スタイルを除去してしまう。
    保存後に xlsx の ZIP 内 sheet XML を走査し、空セルの t="n" 属性を除去することで回避する。
    対象: <c r="..." s="数字" t="n" /> 形式の空セル（子要素なし・値なし）
    """
    import zipfile, shutil, tempfile
    pat = re.compile(r'(<c\s+r="[^"]+"\s+s="\d+"\s+)t="n"\s*(/>)')
    tmp = xlsx_path + '.tmp_fix'
    with zipfile.ZipFile(xlsx_path, 'r') as zin, zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                text = data.decode('utf-8')
                text = pat.sub(r'\1\2', text)
                data = text.encode('utf-8')
            zout.writestr(item, data)
    shutil.move(tmp, xlsx_path)


def _render_layout_to_xlsx(layout: list, grid_params: dict, output_path: str) -> None:
    """
    レイアウト JSON を openpyxl で直接 Excel ファイルに描画する。
    LLM 生成コードを使わずにプログラム的に Excel を生成する。

    COL_OFFSET  = 1  （左1マス余白）
    ROW_PADDING = 1  （ページ上部1マス余白）
    row_offset for page N = (N-1) * max_rows + ROW_PADDING
    page break at page N  = N * (max_rows + ROW_PADDING)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, Alignment, Font
    from openpyxl.worksheet.pagebreak import Break
    from openpyxl.utils import get_column_letter

    COL_OFFSET = 1
    ROW_PADDING = 1
    max_rows = grid_params['max_rows']
    col_width = grid_params.get('excel_col_width', 1.45)
    row_height = grid_params.get('excel_row_height', 11.34)
    paper_size = grid_params.get('paper_size', 9)
    orientation = grid_params.get('orientation', 'portrait')
    default_font_size = grid_params.get('default_font_size', 7)
    font_name     = grid_params.get('font_name', 'MS Gothic')
    margin_left   = grid_params.get('margin_left',   0.43)
    margin_right  = grid_params.get('margin_right',  0.43)
    margin_top    = grid_params.get('margin_top',    0.41)
    margin_bottom = grid_params.get('margin_bottom', 0.41)

    total_pages = len(layout)

    wb = Workbook()
    ws = wb.active
    ws.sheet_format.defaultColWidth = col_width
    ws.sheet_format.defaultRowHeight = row_height
    ws.sheet_format.customHeight = True
    ws.sheet_view.showGridLines = True

    thin = Side(style='thin')

    def _set_border_side(row: int, col: int, **sides) -> None:
        if row < 1:
            return
        try:
            cell = ws.cell(row=row, column=col)
            ex = cell.border
            cell.border = Border(
                top=sides.get('top', ex.top),
                bottom=sides.get('bottom', ex.bottom),
                left=sides.get('left', ex.left),
                right=sides.get('right', ex.right),
            )
        except AttributeError:
            pass

    def _draw_border(s_row: int, e_row: int, s_col: int, e_col: int, borders: dict) -> None:
        # e_row, e_col は exclusive: 枠は s_row <= r < e_row, s_col <= c < e_col
        has_top    = borders.get('top',    True)
        has_bottom = borders.get('bottom', True)
        has_left   = borders.get('left',   True)
        has_right  = borders.get('right',  True)
        if has_top:
            for c in range(s_col, e_col):
                _set_border_side(s_row, c, top=thin)
        if has_bottom:
            for c in range(s_col, e_col):
                _set_border_side(e_row - 1, c, bottom=thin)
        if has_left:
            for r in range(s_row, e_row):
                _set_border_side(r, s_col, left=thin)
        if has_right:
            for r in range(s_row, e_row):
                _set_border_side(r, e_col - 1, right=thin)

    max_used_row = 0
    max_used_col = 0

    for page_layout in layout:
        page_num = page_layout.get('page_number', 1)
        row_offset = (page_num - 1) * max_rows + ROW_PADDING

        for elem in page_layout.get('elements', []):
            etype = elem.get('type')

            if etype == 'text':
                r = elem.get('row', 1) + row_offset
                c = elem.get('col', 1) + COL_OFFSET
                try:
                    cell = ws.cell(row=r, column=c)
                    cell.value = elem.get('content', '')
                    if elem.get('is_vertical'):
                        cell.alignment = Alignment(text_rotation=255, vertical='top', wrap_text=False)
                    else:
                        # [修正] 水平テキストは PDF 座標準拠で left/top を明示。
                        # Excel のデフォルト挙動（数値の自動右揃え等）を上書きし、
                        # PDF の配置位置をそのまま再現する。
                        # セル内改行（\n）を含む場合は wrap_text=True で折り返しを有効化する。
                        cell.alignment = Alignment(
                            horizontal='left', vertical='top',
                            wrap_text=bool(elem.get('multiline')),
                        )
                    # [修正] フォント名: エイリアス解決済みの値、なければグリッドデフォルト
                    resolved_font_name = elem.get('font_name') or font_name
                    # [修正] フォントサイズ: PDF 値を優先しつつ、セル高さに収まる上限でクランプ
                    raw_font_size = elem.get('font_size') or default_font_size
                    max_font_size = row_height * 0.72  # row_height(pt) の約72%を上限とする
                    resolved_font_size = min(float(raw_font_size), max_font_size)
                    font_kwargs = {
                        'name': resolved_font_name,
                        'size': resolved_font_size,
                    }
                    if elem.get('font_color'):
                        font_kwargs['color'] = elem['font_color']
                    cell.font = Font(**font_kwargs)
                except AttributeError:
                    pass
                max_used_row = max(max_used_row, r)
                max_used_col = max(max_used_col, c)

            elif etype == 'border_rect':
                s_row = elem.get('row', 1) + row_offset
                e_row_v = elem.get('end_row', 1) + row_offset
                s_col = elem.get('col', 1) + COL_OFFSET
                e_col_v = elem.get('end_col', 1) + COL_OFFSET
                borders = elem.get('borders', {'top': True, 'bottom': True, 'left': True, 'right': True})
                _draw_border(s_row, e_row_v, s_col, e_col_v, borders)
                max_used_row = max(max_used_row, e_row_v)
                max_used_col = max(max_used_col, e_col_v)

    for pn in range(1, total_pages):
        ws.row_breaks.append(Break(id=pn * (max_rows + ROW_PADDING)))

    ws.page_setup.paperSize = paper_size
    ws.page_setup.orientation = orientation
    ws.page_margins.left   = margin_left
    ws.page_margins.right  = margin_right
    ws.page_margins.top    = margin_top
    ws.page_margins.bottom = margin_bottom

    # print_area はコンテンツ右端に揃える（空白列を含めない）
    # max_cols + COL_OFFSET まで伸ばすと PDF 余白分の空白列が含まれてしまうため、
    # 実際に要素が配置された最右列を使用する。
    if max_used_row > 0 and max_used_col > 0:
        ws.print_area = f"A1:{get_column_letter(max_used_col)}{max_used_row}"

    wb.save(output_path)
    _fix_empty_cell_type_attr(output_path)
    logger.info(f"[render_layout] Excel生成完了: {output_path} ({total_pages} ページ)")


def _generate_border_preview(page_layout: dict, grid_params: dict, output_path: str,
                              pdf_image_path: str | None = None,
                              row_shift: int = 0, col_shift: int = 0) -> None:
    """
    layout の border_rect 要素を PIL キャンバスに描画し、罫線プレビュー画像を生成する。
    pdf_image_path が指定された場合、その画像と同じ解像度・アスペクト比で生成する。

    row_shift / col_shift: auto_layout で除去された余白のセル数。
    プレビュー描画時にこの分だけオフセットを加え、PDF画像と同じ位置に罫線を配置する。
    """
    from PIL import Image, ImageDraw, ImageFont

    max_c = int(grid_params.get('max_cols', 54))
    max_r = int(grid_params.get('max_rows', 42))

    if pdf_image_path and Path(pdf_image_path).exists():
        with Image.open(pdf_image_path) as ref:
            img_w, img_h = ref.size
        cell_w = img_w / max_c
        cell_h = img_h / max_r
    else:
        cell_w = 20.0
        cell_h = 14.0
        img_w = int(cell_w * max_c) + 1
        img_h = int(cell_h * max_r) + 1

    img = Image.new('RGB', (img_w, img_h), 'white')
    draw = ImageDraw.Draw(img)

    def cx(col: float) -> int: return int(col * cell_w)
    def cy(row: float) -> int: return int(row * cell_h)

    for c in range(max_c + 1):
        x = cx(c)
        draw.line([(x, 0), (x, img_h)], fill='#E0E0E0', width=1)
    for r in range(max_r + 1):
        y = cy(r)
        draw.line([(0, y), (img_w, y)], fill='#E0E0E0', width=1)

    border_width = max(2, int(min(cell_w, cell_h) / 7))
    for elem in page_layout.get('elements', []):
        if elem.get('type') != 'border_rect':
            continue
        # シフト量を加算してPDF画像と同じ位置に描画
        r1 = cy(elem['row'] - 1 + row_shift)
        r2 = cy(elem['end_row'] - 1 + row_shift)
        c1 = cx(elem['col'] - 1 + col_shift)
        c2 = cx(elem['end_col'] - 1 + col_shift)
        borders = elem.get('borders', {'top': True, 'bottom': True, 'left': True, 'right': True})
        if borders.get('top',    True): draw.line([(c1, r1), (c2, r1)], fill='black', width=border_width)
        if borders.get('bottom', True): draw.line([(c1, r2), (c2, r2)], fill='black', width=border_width)
        if borders.get('left',   True): draw.line([(c1, r1), (c1, r2)], fill='black', width=border_width)
        if borders.get('right',  True): draw.line([(c2, r1), (c2, r2)], fill='black', width=border_width)

    # コンテンツの有効範囲を計算し、範囲外をグレーアウトする
    border_elems = [e for e in page_layout.get('elements', []) if e.get('type') == 'border_rect']
    if border_elems:
        content_max_col = max(e.get('end_col', e['col']) for e in border_elems)
        content_max_row = max(e.get('end_row', e['row']) for e in border_elems)
        grey_x = cx(content_max_col - 1 + col_shift)
        grey_y = cy(content_max_row - 1 + row_shift)
        grey_fill = (210, 210, 210)
        # コンテンツ右端より右をグレーアウト
        if grey_x < img_w:
            draw.rectangle([(grey_x, 0), (img_w, img_h)], fill=grey_fill)
        # コンテンツ下端より下をグレーアウト
        if grey_y < img_h:
            right_limit = min(grey_x, img_w)
            draw.rectangle([(0, grey_y), (right_limit, img_h)], fill=grey_fill)

    try:
        font = ImageFont.load_default(size=max(8, int(cell_h * 0.8)))
    except TypeError:
        font = ImageFont.load_default()
    label_color = (200, 0, 0)
    # 5セルごとにセル中央にラベルを表示。ラベル番号 = JSON の col/row 値（シフト後）に直接対応。
    # 描画位置はシフト量分ずらして PDF 画像と同じ位置に配置する。
    for c in range(1, max_c + 1, 5):
        lx = cx(c - 1 + col_shift) + cell_w / 2
        if 0 <= lx < img_w:
            draw.text((lx, 1), str(c), fill=label_color, font=font)
    for r in range(1, max_r + 1, 5):
        ly = cy(r - 1 + row_shift) + cell_h / 2
        if 0 <= ly < img_h:
            draw.text((1, ly), str(r), fill=label_color, font=font)

    img.save(output_path)



def _has_japanese(text: str) -> bool:
    """文字列に日本語文字（漢字・ひらがな・カタカナ・全角記号）が含まれるか判定する。"""
    return any(
        '\u3040' <= c <= '\u30ff'  # ひらがな・カタカナ
        or '\u4e00' <= c <= '\u9fff'  # CJK 統合漢字
        or '\uff00' <= c <= '\uffef'  # 全角英数・記号
        for c in text
    )


def _join_word_texts(texts: list) -> str:
    """
    word テキストのリストを結合する。
    テキスト結合ルール:
      - 日本語文字を含む場合はスペースなし
      - 英数字のみの場合は半角スペースで結合
    """
    combined = ''.join(texts)
    if _has_japanese(combined):
        return combined
    return ' '.join(t for t in texts if t.strip())


def _split_by_horizontal_gap(words: list, gap_factor: float = 2.0) -> list:
    """
    水平方向のギャップでワードリストを分割する。
    前のワードの x1 と次のワードの x0 の間隔がフォントサイズ × gap_factor を
    超えた場合、別グループとして分割する。
    返り値: ワードリストのリスト（各サブリストは水平に連続するワード群）
    """
    if len(words) <= 1:
        return [words]
    sw = sorted(words, key=lambda w: float(w.get('x0', 0)))
    groups: list = [[sw[0]]]
    for w in sw[1:]:
        prev = groups[-1][-1]
        prev_x1 = float(prev.get('x1', prev.get('x0', 0)))
        curr_x0 = float(w.get('x0', 0))
        gap = curr_x0 - prev_x1
        # 閾値: 前ワードと現ワードのフォントサイズの平均 × gap_factor
        avg_fs = (float(prev.get('font_size', prev.get('size', 10)))
                  + float(w.get('font_size', w.get('size', 10)))) / 2
        threshold = avg_fs * gap_factor
        if gap > threshold:
            groups.append([w])
        else:
            groups[-1].append(w)
    return groups


def _table_text_elements_from_2d(page: dict, grid_params: dict) -> list:
    """
    pdfplumber の extract_tables() が返す 2D 配列と列/行境界座標を使って、
    テーブル内テキスト要素を生成する。

    - None はセル結合の延長を意味するのでスキップ
    - 連続 None から列スパン・行スパンを検出し end_col を算出
    - '\\n' 区切りの複数行は行スパン内に分散配置
    """
    max_rows = grid_params['max_rows']
    max_cols = grid_params['max_cols']

    grid_h = float(page['height']) / max_rows
    grid_w = float(page['width'])  / max_cols

    def to_row(y: float) -> int:
        return max(1, min(max_rows, 1 + int(float(y) / grid_h)))

    def to_col(x: float) -> int:
        return max(1, min(max_cols, 1 + int(float(x) / grid_w)))

    row_shift = page.get('_row_shift', 0)
    col_shift = page.get('_col_shift', 0)

    elements: list = []
    # 同一 word が複数セルの bbox に含まれる場合の重複処理を防止
    _used_word_ids: set = set()

    # table_data_raw は \n 保持版（複数行検出に使用）。なければ cleaned 版にフォールバック
    _table_data_src = page.get('table_data_raw') or page.get('table_data', [])

    for table_data, col_xs, row_ys, cells_2d in zip(
        _table_data_src,
        page.get('table_col_x_positions', []),
        page.get('table_row_y_positions', []),
        page.get('table_cells', []),
    ):
        if not table_data or not col_xs or not row_ys:
            continue

        num_rows = len(table_data)
        num_cols = len(table_data[0]) if table_data else 0

        for r_idx, trow in enumerate(table_data):
            if r_idx >= len(row_ys) - 1:
                continue

            for c_idx, cell_content in enumerate(trow):
                # None = 結合セルの延長なのでスキップ
                if cell_content is None:
                    continue

                raw = cell_content if isinstance(cell_content, str) else ''
                lines = [ln.strip() for ln in raw.split('\n') if ln.strip()]
                if not lines:
                    continue

                if c_idx >= len(col_xs) - 1:
                    continue

                # table.rows から直接セル bbox を取得（結合セル対応の正確な座標）。
                # cells_2d[r_idx][c_idx] は None（結合延長）でないことが保証されているが、
                # 次元不整合に備えてフォールバックを設ける。
                cell_bbox = None
                if (cells_2d
                        and r_idx < len(cells_2d)
                        and c_idx < len(cells_2d[r_idx])):
                    cell_bbox = cells_2d[r_idx][c_idx]

                if cell_bbox is not None:
                    # 直接セル bbox から座標を取得
                    # 縦結合セルでも x0/x1 は正確。y0 はそのセルの開始行トップ。
                    x0 = cell_bbox['x0']
                    x1 = cell_bbox['x1']
                    y0 = cell_bbox['top']
                    y1 = cell_bbox['bottom']
                else:
                    # フォールバック: 旧来の sorted unique 座標インデックス方式
                    col_end_idx = c_idx + 1
                    while col_end_idx < num_cols and trow[col_end_idx] is None:
                        col_end_idx += 1
                    x0 = col_xs[c_idx]
                    x1 = col_xs[min(col_end_idx, len(col_xs) - 1)]
                    y0 = row_ys[r_idx]
                    y1 = row_ys[min(r_idx + 1, len(row_ys) - 1)]

                # シフト適用済みグリッド座標
                grid_row = max(1, to_row(y0) - row_shift)
                grid_col = max(1, to_col(x0) - col_shift)
                grid_end_col = max(grid_col + 1, min(max_cols, to_col(x1) - col_shift))

                # セル内の words を実座標ベースで検索し、word 座標で配置する。
                # word が見つからない場合のみ 2D テキストのフォールバックを使用する。
                # 既に別セルで処理済みの word は除外する。
                cell_words: list = []
                for w in page.get('words', []):
                    if '_row' not in w:
                        continue
                    wid = id(w)
                    if wid in _used_word_ids:
                        continue
                    wx0 = float(w.get('x0', 0))
                    wy0 = float(w.get('top', 0))
                    if (x0 - 2.0 <= wx0 <= x1 + 2.0
                            and y0 - 2.0 <= wy0 <= y1 + 2.0):
                        cell_words.append(w)

                if cell_words:
                    # 処理済みとしてマーク
                    for w in cell_words:
                        _used_word_ids.add(id(w))
                    # word 座標ベース配置: _row でグループ化
                    # テキストがセルの border_rect 外にはみ出さないようクリップ
                    cell_max_row = max(grid_row, to_row(y1) - row_shift - 1)
                    cell_word_rows: dict = {}
                    for w in cell_words:
                        wr_clipped = min(w['_row'], cell_max_row)
                        cell_word_rows.setdefault(wr_clipped, []).append(w)

                    for wr, wds in sorted(cell_word_rows.items()):
                        # 同一 _row 内でも top 座標のギャップで視覚行分割する。
                        _VIS_GAP = 3.0
                        sw_cell = sorted(wds, key=lambda w: float(w.get('top', 0)))
                        vis_lines_cell: list = [[sw_cell[0]]]
                        for _cw in sw_cell[1:]:
                            prev_bottom = max(float(v.get('bottom', v.get('top', 0))) for v in vis_lines_cell[-1])
                            this_top = float(_cw.get('top', 0))
                            if this_top - prev_bottom > _VIS_GAP:
                                vis_lines_cell.append([_cw])
                            else:
                                vis_lines_cell[-1].append(_cw)

                        for vl_idx, vl_words in enumerate(vis_lines_cell):
                            vl_row = wr + vl_idx
                            # 水平ギャップで分割して別要素として配置
                            h_groups = _split_by_horizontal_gap(
                                sorted(vl_words, key=lambda x: float(x.get('x0', 0)))
                            )
                            for hg in h_groups:
                                line_text = _join_word_texts(
                                    [w.get('text', '') for w in hg]
                                ).strip()
                                if not line_text:
                                    continue
                                hg_col = max(1, to_col(float(hg[0].get('x0', 0))) - col_shift)
                                hg_end_col = max(hg_col + 1, min(max_cols, to_col(float(hg[-1].get('x1', hg[-1].get('x0', 0)))) - col_shift))
                                first_w = hg[0]
                                te: dict = {
                                    'type': 'text',
                                    'content': line_text,
                                    'row': min(max_rows, vl_row),
                                    'col': hg_col,
                                    'end_col': hg_end_col,
                                }
                                if first_w.get('font_color') and first_w['font_color'] != '000000':
                                    te['font_color'] = first_w['font_color']
                                if first_w.get('font_size'):
                                    te['font_size'] = first_w['font_size']
                                _fn = _normalize_font_name(first_w.get('fontname', ''))
                                if _fn:
                                    te['font_name'] = _fn
                                elements.append(te)
                    continue  # word ベース配置完了

                # フォールバック: words が見つからない場合は 2D テキストを分散配置
                grid_end_row = max(grid_row, to_row(y1) - row_shift - 1)
                for line_idx, line in enumerate(lines):
                    line_row = grid_row + line_idx
                    if line_row > grid_end_row:
                        break
                    elements.append({
                        'type': 'text',
                        'content': line,
                        'row': min(max_rows, line_row),
                        'col': grid_col,
                        'end_col': grid_end_col,
                    })

    return elements


def _fill_missing_text(layout_json_str: str, extracted_data: dict, grid_params: dict | None = None) -> str:
    """
    LLMが生成したレイアウトJSONに対し、extracted_dataのwordsと照合して
    欠落しているテキスト要素をプログラム的に補完する。

    Step 1 / Step 1.5 の LLM が見落とした word を確実に補う。
    既に text 要素が存在する (row, col) には追加しない（上書き禁止）。
    """
    try:
        layout = json.loads(layout_json_str)
    except (json.JSONDecodeError, ValueError):
        return layout_json_str  # パース失敗時はそのまま返す

    total_added = 0
    for page_layout in layout:
        page_num = page_layout.get('page_number', 1)
        page_data = next(
            (p for p in extracted_data['pages'] if p['page_number'] == page_num),
            None,
        )
        if not page_data:
            continue

        # 既存 text 要素の (row, col) を収集
        existing: set = set()
        for elem in page_layout.get('elements', []):
            if elem.get('type') == 'text' and 'row' in elem and 'col' in elem:
                existing.add((elem['row'], elem['col']))

        # words を (_row, _col) でグループ化
        # テーブル内ワードは _auto_generate_layout で 2D 配列から処理済みのためスキップ
        # ただしテーブル2D配列でNoneセルに該当するワードはテーブル外として扱う
        _tol_f = 2.0
        _tbboxes = page_data.get('table_bboxes', [])

        _tcell_bboxes: list = []
        _td_src = page_data.get('table_data_raw') or page_data.get('table_data', [])
        for _td, _c2d in zip(_td_src, page_data.get('table_cells', [])):
            if not _td or not _c2d:
                continue
            for _ri, _trow in enumerate(_td):
                for _ci, _cc in enumerate(_trow):
                    if _cc is None:
                        continue
                    if (_c2d and _ri < len(_c2d)
                            and _ci < len(_c2d[_ri])
                            and _c2d[_ri][_ci] is not None):
                        _cb = _c2d[_ri][_ci]
                        _tcell_bboxes.append(
                            (float(_cb['x0']), float(_cb['top']),
                             float(_cb['x1']), float(_cb['bottom']))
                        )

        def _in_tbl(w: dict) -> bool:
            wx = float(w.get('x0', 0))
            wy = float(w.get('top', 0))
            in_any = False
            for _b in _tbboxes:
                if (_b[0] - _tol_f <= wx <= _b[2] + _tol_f and
                        _b[1] - _tol_f <= wy <= _b[3] + _tol_f):
                    in_any = True
                    break
            if not in_any:
                return False
            for _cb in _tcell_bboxes:
                if (_cb[0] - _tol_f <= wx <= _cb[2] + _tol_f and
                        _cb[1] - _tol_f <= wy <= _cb[3] + _tol_f):
                    return True
            return False

        groups: dict = {}
        for w in page_data.get('words', []):
            if '_row' not in w or '_col' not in w:
                continue
            if _in_tbl(w):
                continue
            key = (w['_row'], w['_col'])
            groups.setdefault(key, []).append(w)

        added = []
        for (row, col), words in sorted(groups.items()):
            # 水平ギャップで分割して2段表示を検出
            h_groups = _split_by_horizontal_gap(words)
            for hg in h_groups:
                hg_col = hg[0].get('_col', col)
                if (row, hg_col) in existing:
                    continue
                content = _join_word_texts([w.get('text', '') for w in hg])
                stripped = content.strip()
                # 空白・純粋な区切り記号（ASCII句読点の1文字）はスキップ
                # ただし △▼○● 等の図形記号・日本語1文字は意味があるため残す
                if not stripped or (len(stripped) == 1 and stripped in '!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~'):
                    continue
                last_w = hg[-1]
                if grid_params:
                    _grid_w = float(page_data['width']) / grid_params['max_cols']
                    _max_cols = grid_params['max_cols']
                    hg_end = max(hg_col + 1, min(_max_cols, 1 + int(float(last_w.get('x1', last_w.get('x0', 0))) / _grid_w)))
                else:
                    hg_end = hg_col + len(content)
                elem: dict = {
                    'type': 'text',
                    'content': content,
                    'row': row,
                    'col': hg_col,
                    'end_col': hg_end,
                }
                first = hg[0]
                if first.get('font_color') and first['font_color'] != '000000':
                    elem['font_color'] = first['font_color']
                if first.get('font_size'):
                    elem['font_size'] = first['font_size']
                font_name = _normalize_font_name(first.get('fontname', ''))
                if font_name:
                    elem['font_name'] = font_name
                added.append(elem)

        if added:
            page_layout['elements'].extend(added)
            total_added += len(added)

    if total_added:
        logger.info(f"[fill_missing_text] {total_added} 個の欠落テキスト要素を補完しました")

    return json.dumps(layout, ensure_ascii=False)


def _auto_generate_layout(extracted_data: dict, grid_params: dict) -> str:
    """
    extracted_data から直接レイアウトJSONを生成する（step1 + step1.5 のスクリプト代替）。

    - table_border_rects / rects → border_rect 要素
    - words を (_row, _col) でグループ化 → text 要素
    - 座標整合性チェック・重複除去・クランプを適用
    """
    max_rows = grid_params['max_rows']
    max_cols = grid_params['max_cols']

    def _is_near_duplicate(seen: list, r: int, er: int, c: int, ec: int, tol: int = 1) -> bool:
        """グリッド座標が tol 以内の既登録要素があれば重複とみなす。
        完全一致チェックだけでは grid 量子化誤差（±1行/列ずれ）による
        実質同一矩形の二重登録を取りこぼすため、近似一致も除外する。"""
        for (sr, ser, sc, sec) in seen:
            if (abs(sr - r) <= tol and abs(ser - er) <= tol and
                    abs(sc - c) <= tol and abs(sec - ec) <= tol):
                return True
        return False

    layout = []
    for page in extracted_data.get('pages', []):
        elements = []
        seen_border_rects: list = []  # [修正] set → list（近似一致検索のため）
        grid_w = float(page['width']) / max_cols

        # table_border_rects → border_rect 要素
        for tbr in page.get('table_border_rects', []):
            r  = min(tbr['_row'],     max_rows)
            er = min(tbr['_end_row'], max_rows)
            c  = min(tbr['_col'],     max_cols)
            ec = min(tbr['_end_col'], max_cols)
            if r > er: r, er = er, r
            if c > ec: c, ec = ec, c
            if r == er and c == ec:
                continue
            # [修正] 完全一致 → ±1グリッド近似一致で重複を除外
            if _is_near_duplicate(seen_border_rects, r, er, c, ec):
                continue
            seen_border_rects.append((r, er, c, ec))
            elements.append({
                'type': 'border_rect',
                'row': r, 'end_row': er,
                'col': c, 'end_col': ec,
                'borders': tbr.get('_borders', {'top': True, 'bottom': True, 'left': True, 'right': True}),
            })

        # rects → border_rect 要素
        # 薄い矩形（線の太さ分）は水平線/垂直線として処理する。
        for rect in page.get('rects', []):
            if '_row' not in rect:
                continue
            r  = min(rect['_row'],     max_rows)
            er = min(rect['_end_row'], max_rows)
            c  = min(rect['_col'],     max_cols)
            ec = min(rect['_end_col'], max_cols)
            if r > er: r, er = er, r
            if c > ec: c, ec = ec, c

            # 横線 (r == er): 水平罫線として描画
            if r == er and c != ec:
                key = (r, r + 1, c, ec)
                if _is_near_duplicate(seen_border_rects, *key):
                    continue
                seen_border_rects.append(key)
                elements.append({
                    'type': 'border_rect',
                    'row': r, 'end_row': r + 1,
                    'col': c, 'end_col': ec,
                    'borders': {'top': True, 'bottom': False, 'left': False, 'right': False},
                })
                continue

            # 縦線 (c == ec): 垂直罫線として描画
            if c == ec and r != er:
                key = (r, er, c, c + 1)
                if _is_near_duplicate(seen_border_rects, *key):
                    continue
                seen_border_rects.append(key)
                elements.append({
                    'type': 'border_rect',
                    'row': r, 'end_row': er,
                    'col': c, 'end_col': c + 1,
                    'borders': {'top': False, 'bottom': False, 'left': True, 'right': False},
                })
                continue

            # 通常の矩形
            if r == er and c == ec:
                continue
            if _is_near_duplicate(seen_border_rects, r, er, c, ec):
                continue
            seen_border_rects.append((r, er, c, ec))
            elements.append({
                'type': 'border_rect',
                'row': r, 'end_row': er,
                'col': c, 'end_col': ec,
                'borders': {'top': True, 'bottom': True, 'left': True, 'right': True},
            })

        # words → text 要素（_row, _col でグループ化）
        # テーブル内ワードは _table_text_elements_from_2d で処理するためスキップする。
        # ただしテーブル2D配列でNoneセル（結合延長）に該当するワードは
        # テーブルで処理されないため、テーブル外として扱う。
        _tol = 2.0
        _table_bboxes = page.get('table_bboxes', [])

        # テーブルの有効セル（None でない）のbboxリストを構築
        _table_cell_bboxes: list = []
        _table_data_src = page.get('table_data_raw') or page.get('table_data', [])
        for _td, _cells_2d in zip(
            _table_data_src,
            page.get('table_cells', []),
        ):
            if not _td or not _cells_2d:
                continue
            for _ri, _trow in enumerate(_td):
                for _ci, _cell_content in enumerate(_trow):
                    if _cell_content is None:
                        continue  # 結合延長セル → ワードはここには属さない
                    if (_cells_2d and _ri < len(_cells_2d)
                            and _ci < len(_cells_2d[_ri])
                            and _cells_2d[_ri][_ci] is not None):
                        cb = _cells_2d[_ri][_ci]
                        _table_cell_bboxes.append(
                            (float(cb['x0']), float(cb['top']),
                             float(cb['x1']), float(cb['bottom']))
                        )

        def _in_table(w: dict) -> bool:
            wx = float(w.get('x0', 0))
            wy = float(w.get('top', 0))
            # まずテーブルbbox内かチェック
            in_any_table = False
            for _bbox in _table_bboxes:
                if (_bbox[0] - _tol <= wx <= _bbox[2] + _tol and
                        _bbox[1] - _tol <= wy <= _bbox[3] + _tol):
                    in_any_table = True
                    break
            if not in_any_table:
                return False
            # テーブルbbox内でも、有効セルのbboxに含まれているか確認
            for cb in _table_cell_bboxes:
                if (cb[0] - _tol <= wx <= cb[2] + _tol and
                        cb[1] - _tol <= wy <= cb[3] + _tol):
                    return True
            return False  # テーブルbbox内だが有効セルに属さない → テーブル外扱い

        groups: dict = {}
        for w in page.get('words', []):
            if '_row' not in w or '_col' not in w:
                continue
            if _in_table(w):
                continue  # テーブル内は 2D 配列から生成
            key = (w['_row'], w['_col'])
            groups.setdefault(key, []).append(w)

        seen_text: set = set()
        for (row, col), words in sorted(groups.items()):
            # 同一グループ内に複数の視覚行が含まれる場合（_SPLIT_GAP 以内の小さなギャップ）、
            # 行ごとに \n で結合してセル内改行として表現する。
            _INLINE_LINE_GAP = 1.0  # pt: この値を超えるギャップを行区切りとみなす
            # PDF が同一座標に重複ワードを出力する場合（影付きテキスト等）を除去する。
            # (text, top×0.5pt丸め, x0×0.5pt丸め) が同一のワードは重複とみなす。
            _seen_w: set = set()
            _deduped: list = []
            for _w in words:
                _wk = (_w.get('text', ''),
                       round(float(_w.get('top', 0)) * 2) / 2,
                       round(float(_w.get('x0', 0)) * 2) / 2)
                if _wk not in _seen_w:
                    _seen_w.add(_wk)
                    _deduped.append(_w)
            words = _deduped
            sw = sorted(words, key=lambda w: float(w.get('top', 0)))
            vis_lines: list = [[sw[0]]]
            for _w in sw[1:]:
                prev_b = float(vis_lines[-1][-1].get('bottom', vis_lines[-1][-1]['top']))
                this_t = float(_w.get('top', 0))
                if this_t - prev_b > _INLINE_LINE_GAP:
                    vis_lines.append([_w])
                else:
                    vis_lines[-1].append(_w)
            row_c = min(row, max_rows)
            col_c = min(col, max_cols)

            if len(vis_lines) > 1:
                # 複数の視覚行がある場合は各行を別々の text 要素として配置する。
                # \n 結合 + wrap_text=True では行高さが固定のため2行目以降が見えないため。
                # 各行の先頭ワードの _row（シフト済みグリッド行）を使い、
                # 最低でも1行ずつ下にずれるよう保証する。
                prev_row_c = row_c - 1
                for _line in vis_lines:
                    # 水平ギャップで分割して2段表示を検出
                    _h_groups = _split_by_horizontal_gap(_line)
                    _word_row = _line[0].get('_row', row_c)
                    _line_row_c = min(max_rows, max(prev_row_c + 1, _word_row))
                    for _hg in _h_groups:
                        _line_content = _join_word_texts([_w.get('text', '') for _w in _hg])
                        _stripped = _line_content.strip()
                        if not _stripped or (len(_stripped) == 1 and _stripped in '!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~'):
                            continue
                        _hg_col = _hg[0].get('_col', col_c)
                        _hg_col_c = min(_hg_col, max_cols)
                        _pos = (_line_row_c, _hg_col_c)
                        if _pos in seen_text:
                            continue
                        seen_text.add(_pos)
                        first = _hg[0]
                        last = _hg[-1]
                        _hg_end_col = max(_hg_col_c + 1, min(max_cols, 1 + int(float(last.get('x1', last.get('x0', 0))) / grid_w)))
                        _elem: dict = {
                            'type': 'text',
                            'content': _line_content,
                            'row': _line_row_c,
                            'col': _hg_col_c,
                            'end_col': _hg_end_col,
                        }
                        if first.get('font_color') and first['font_color'] != '000000':
                            _elem['font_color'] = first['font_color']
                        if first.get('font_size'):
                            _elem['font_size'] = first['font_size']
                        _fn = _normalize_font_name(first.get('fontname', ''))
                        if _fn:
                            _elem['font_name'] = _fn
                        elements.append(_elem)
                    prev_row_c = _line_row_c
                continue  # vis_lines > 1 の場合はここで処理完了

            # 単一視覚行でも水平ギャップ分割を適用
            h_groups = _split_by_horizontal_gap(words)
            for hg in h_groups:
                content = _join_word_texts([w.get('text', '') for w in hg])
                stripped = content.strip()
                if not stripped or (len(stripped) == 1 and stripped in '!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~'):
                    continue
                hg_col = hg[0].get('_col', col_c)
                hg_col_c = min(hg_col, max_cols)
                pos = (row_c, hg_col_c)
                if pos in seen_text:
                    continue
                seen_text.add(pos)
                first = hg[0]
                last_w = hg[-1]
                hg_end_col = max(hg_col_c + 1, min(max_cols, 1 + int(float(last_w.get('x1', last_w.get('x0', 0))) / grid_w)))
                elem: dict = {
                    'type': 'text',
                    'content': content,
                    'row': row_c,
                    'col': hg_col_c,
                    'end_col': hg_end_col,
                }
                if first.get('font_color') and first['font_color'] != '000000':
                    elem['font_color'] = first['font_color']
                if first.get('font_size'):
                    elem['font_size'] = first['font_size']
                font_name = _normalize_font_name(first.get('fontname', ''))
                if font_name:
                    elem['font_name'] = font_name
                if first.get('is_vertical'):
                    elem['is_vertical'] = True
                    if '_end_row' in first:
                        elem['end_row'] = min(first['_end_row'], max_rows)
                    elem['end_col'] = min(hg_col_c + 1, max_cols)
                elements.append(elem)

        # テーブル内テキストを 2D 配列から生成（merged cell / colspan 対応）
        for tbl_elem in _table_text_elements_from_2d(page, grid_params):
            pos = (tbl_elem['row'], tbl_elem['col'])
            if pos not in seen_text:
                seen_text.add(pos)
                elements.append(tbl_elem)

        layout.append({'page_number': page['page_number'], 'elements': elements})

    return json.dumps(layout, ensure_ascii=False)


def _setup_grid_params(first_page: dict, grid_size: str) -> dict:
    """
    グリッドパラメータを設定する。
    A4固定の前提で GRID_SIZES の値をそのまま使用する（動的スケーリング不要）。
    横向き(landscape)の場合は max_cols_landscape / max_rows_landscape を使用する。
    """
    ref = GRID_SIZES.get(grid_size, GRID_SIZES["1pt"])
    grid_params = dict(ref)
    grid_params['grid_size'] = grid_size

    # 用紙サイズ・向き検出
    max_dim_pt = max(first_page['width'], first_page['height'])
    grid_params['paper_size'] = 8 if max_dim_pt > 1000 else 9  # 8=A3, 9=A4
    is_landscape = first_page['width'] > first_page['height']
    grid_params['orientation'] = 'landscape' if is_landscape else 'portrait'

    # 横向きの場合は専用の max_cols / max_rows を上書き
    if is_landscape:
        grid_params['max_cols'] = ref['max_cols_landscape']
        grid_params['max_rows'] = ref['max_rows_landscape']

    logger.debug(
        f"[grid] {grid_size} ({grid_params['orientation']}): "
        f"max_cols={grid_params['max_cols']}, max_rows={grid_params['max_rows']}, "
        f"excel_col_width={grid_params['excel_col_width']}"
    )

    return grid_params


class SheetlingPipeline:
    """PDF から Excel 方眼紙を自動生成するパイプライン。"""

    def __init__(self, output_base_dir: str):
        self.output_base_dir = Path(output_base_dir)

    def auto_layout(self, pdf_path: str, in_base_dir: str = "data/in", grid_size: str = "small") -> dict:
        """
        [全自動] PDF → Excel 高精度解析パイプライン (Sheetling-pre 方式)。

        1. extract_pdf_data() でテキスト・罫線を抽出
        2. _setup_grid_params() で A4 ポイント基準の方眼密度を動的計算
        3. _compute_grid_coords() で全要素にグリッド座標を付与
        4. _merge_table_border_rects() で結合セルを復元 (Pre版ロジック)
        5. 行シフト補正で上部余白を詰め、コンテンツを上詰めに配置
        6. _auto_generate_layout() でプログラム的にレイアウト JSON を生成
        7. _fill_missing_text() で欠落したテキスト要素を最終補完
        8. _render_layout_to_xlsx() で Excel を直接レンダリング
        """
        logger.info(f"--- [auto] PDF → Excel 高精度自動生成: {Path(pdf_path).name} ---")
        path_obj = Path(pdf_path)
        pdf_name = path_obj.stem

        try:
            rel_path = path_obj.parent.relative_to(Path(in_base_dir))
            out_dir = self.output_base_dir / rel_path
        except ValueError:
            out_dir = self.output_base_dir / pdf_name

        out_dir.mkdir(parents=True, exist_ok=True)
        prompts_dir = out_dir / "prompts" / grid_size
        prompts_dir.mkdir(parents=True, exist_ok=True)

        layout_json_name = f"{pdf_name}_{grid_size}_layout.json"
        grid_params_name = f"{pdf_name}_{grid_size}_grid_params.json"

        # PDF抽出 & グリッド座標付与
        extracted_data = extract_pdf_data(pdf_path)
        first_page = extracted_data['pages'][0]
        grid_params = _setup_grid_params(first_page, grid_size)
        for page in extracted_data['pages']:
            _compute_grid_coords(page, grid_params['max_rows'], grid_params['max_cols'])

        # ---- [Sheetling-pre 移植ロジック] -----------------------------------------
        # 結合セルのマージは不要（有効セルベースで生成済み）

        # 行シフト補正（上部余白の除去）
        for page in extracted_data['pages']:
            all_rows = (
                [w['_row'] for w in page['words'] if '_row' in w]
                + [r['_row'] for r in page['rects'] if '_row' in r]
                + [tbr['_row'] for tbr in page['table_border_rects']]
            )
            row_shift = (min(all_rows) - 1) if all_rows else 0
            page['_row_shift'] = row_shift  # _table_text_elements_from_2d で使用
            if row_shift > 0:
                for w in page['words']:
                    if '_row' in w:
                        w['_row'] -= row_shift
                        if '_end_row' in w: w['_end_row'] -= row_shift
                for r in page['rects']:
                    if '_row' in r:
                        r['_row'] -= row_shift
                        r['_end_row'] -= row_shift
                for tbr in page['table_border_rects']:
                    tbr['_row'] -= row_shift
                    tbr['_end_row'] -= row_shift

        # 3. 列シフト補正（左余白を1列に統一）
        # 行シフト補正と同様に、コンテンツ最左列を1にそろえ、
        # COL_OFFSET=1 で左側ちょうど1列の余白になるようにする。
        for page in extracted_data['pages']:
            all_cols = (
                [w['_col'] for w in page['words'] if '_col' in w]
                + [r['_col'] for r in page['rects'] if '_col' in r]
                + [tbr['_col'] for tbr in page['table_border_rects']]
            )
            col_shift = (min(all_cols) - 1) if all_cols else 0
            page['_col_shift'] = col_shift  # _table_text_elements_from_2d で使用
            if col_shift > 0:
                for w in page['words']:
                    if '_col' in w:
                        w['_col'] -= col_shift
                for r in page['rects']:
                    if '_col' in r:
                        r['_col'] -= col_shift
                        r['_end_col'] -= col_shift
                for tbr in page['table_border_rects']:
                    tbr['_col'] -= col_shift
                    tbr['_end_col'] -= col_shift
        # ---------------------------------------------------------------------------

        # デバッグ用に中間データを保存
        with open(out_dir / f"{pdf_name}_extracted.json", "w", encoding="utf-8") as f:
            json.dump(extracted_data, f, indent=2, ensure_ascii=False)
        with open(out_dir / grid_params_name, "w", encoding="utf-8") as f:
            json.dump(grid_params, f, ensure_ascii=False)

        # レイアウトJSON生成 & 欠落テキスト補完
        layout_json_str = _auto_generate_layout(extracted_data, grid_params)
        filled_json_str = _fill_missing_text(layout_json_str, extracted_data, grid_params)
        layout_data = json.loads(filled_json_str)

        # プレビュー生成用にページごとのシフト量を退避
        _page_shifts = {}
        for page in extracted_data['pages']:
            _pn = page.get('page_number', 1)
            _page_shifts[_pn] = {
                'row_shift': page.get('_row_shift', 0),
                'col_shift': page.get('_col_shift', 0),
            }

        # table_data / table_row_y_positions / table_cells は layout 生成後不要なため削除
        for page in extracted_data['pages']:
            page.pop('table_data', None)
            page.pop('table_data_raw', None)
            page.pop('table_row_y_positions', None)
            page.pop('table_cells', None)
            page.pop('_row_shift', None)
            page.pop('_col_shift', None)

        # アーカイブ・correct コマンド用保存
        output_json_path = out_dir / layout_json_name
        with open(output_json_path, "w", encoding="utf-8") as f:
            f.write(filled_json_str)

        # 直接 Excel レンダリング (Pre版方式)
        # _1pt / _2pt のみサフィックスを付与。それ以外は従来どおり。
        _xlsx_suffix = f"_{grid_size}" if grid_size in ("1pt", "2pt") else ""
        xlsx_path = out_dir / f"{pdf_name}_Python版{_xlsx_suffix}.xlsx"
        try:
            _render_layout_to_xlsx(layout_data, grid_params, str(xlsx_path))
            logger.info(f"✅ Excel 生成完了: {xlsx_path.name}")
        except Exception as e:
            logger.error(f"❌ Excel 生成に失敗しました: {e}")
            raise

        # ---- ビジョンレビュー素材の自動生成 ----------------------------------------
        # correct コマンドで AI 視覚比較できるよう、PDF 画像・罫線プレビュー・
        # VISUAL_REVIEW_PROMPT・corrections テンプレートをまとめて出力する。
        # prompts_dir はすでに auto_layout() 冒頭で作成済み。
        try:
            import pdfplumber as _pdfplumber
            with _pdfplumber.open(pdf_path) as _pdf:
                for _pg in _pdf.pages:
                    _pn = _pg.page_number
                    _pdir = prompts_dir / f"page_{_pn}"
                    _pdir.mkdir(parents=True, exist_ok=True)
                    _img = _pg.to_image(resolution=144)
                    _img.save(str(_pdir / f"{pdf_name}_page{_pn}.png"))
            logger.info(f"  PDF ページ画像を生成しました: {prompts_dir}/page_N/")
        except Exception as _e:
            logger.warning(f"PDF ページ画像の生成に失敗しました（correct フェーズは利用不可）: {_e}")

        for _page_layout in layout_data:
            _pn = _page_layout.get('page_number', 1)
            _pdir = prompts_dir / f"page_{_pn}"
            _pdir.mkdir(parents=True, exist_ok=True)

            _pdf_img = _pdir / f"{pdf_name}_page{_pn}.png"
            _preview  = _pdir / f"{pdf_name}_excel_page{_pn}.png"
            _shifts = _page_shifts.get(_pn, {'row_shift': 0, 'col_shift': 0})
            try:
                _generate_border_preview(_page_layout, grid_params, str(_preview),
                                         pdf_image_path=str(_pdf_img),
                                         row_shift=_shifts['row_shift'],
                                         col_shift=_shifts['col_shift'])
            except Exception as _e:
                logger.warning(f"  ページ {_pn}: 罫線プレビュー生成に失敗しました: {_e}")

            _gp_for_prompt = dict(grid_params)
            _gp_for_prompt.setdefault('position_tolerance_cells', '1〜2')
            # コンテンツの有効範囲を計算（AIが範囲外の座標を指定しないよう制約する）
            _elems = _page_layout.get('elements', [])
            _all_end_rows = [e.get('end_row', e.get('row', 1)) for e in _elems if e.get('type') == 'border_rect']
            _all_end_cols = [e.get('end_col', e.get('col', 1)) for e in _elems if e.get('type') == 'border_rect']
            _gp_for_prompt['content_max_row'] = max(_all_end_rows) if _all_end_rows else grid_params['max_rows']
            _gp_for_prompt['content_max_col'] = max(_all_end_cols) if _all_end_cols else grid_params['max_cols']
            _prompt_text = VISUAL_REVIEW_PROMPT.format(page_number=_pn, **_gp_for_prompt)
            (_pdir / f"{pdf_name}_visual_review_page{_pn}.txt").write_text(_prompt_text, encoding="utf-8")

            _corr_path = _pdir / f"{pdf_name}_visual_corrections_page{_pn}.json"
            if not _corr_path.exists():
                _corr_path.write_text('{"corrections": []}', encoding="utf-8")

        logger.info(
            f"  [review 素材] prompts/{grid_size}/page_N/ に PDF 画像・罫線プレビュー・プロンプトを出力しました\n"
            f"  次のステップ:\n"
            f"    1. PDF 画像と罫線プレビューを AI に渡し、visual_review プロンプトで比較させる\n"
            f"    2. AI の出力 JSON を visual_corrections_page{{N}}.json に保存する\n"
            f"    3. python -m src.main correct --pdf {pdf_name} --grid-size {grid_size} を実行する"
        )
        # -------------------------------------------------------------------------

        return {
            "xlsx_path": str(xlsx_path),
            "layout_json": str(output_json_path),
            "grid_params": grid_params
        }


    def apply_corrections(self, pdf_name: str, corrections_json: str, specific_out_dir: str = None,
                          layout_json_name: str = None) -> None:
        """
        ビジョンLLMが出力した修正指示を _layout.json に適用する。

        corrections_json の形式:
        {
          "corrections": [
            {"action": "add_text",    "page": 1, "row": 3, "col": 5, "content": "テキスト"},
            {"action": "fix_text",    "page": 1, "row": 3, "col": 5, "new_row": 4, "new_col": 6},
            {"action": "add_border",  "page": 1, "row": 3, "end_row": 8, "col": 2, "end_col": 15,
                                      "borders": {"top": true, ...}},
            {"action": "remove_border","page": 1, "row": 3, "end_row": 8, "col": 2, "end_col": 15}
          ]
        }
        """
        if specific_out_dir:
            out_dir = Path(specific_out_dir)
        else:
            out_dir = self.output_base_dir / pdf_name

        _layout_json_name = layout_json_name or f"{pdf_name}_layout.json"

        output_json_path = out_dir / _layout_json_name

        if not output_json_path.exists():
            raise FileNotFoundError(f"_layout.json が見つかりません: {output_json_path}")

        layout = json.loads(output_json_path.read_text(encoding="utf-8"))

        try:
            corrections_data = json.loads(corrections_json)
            corrections = corrections_data.get("corrections", [])
        except json.JSONDecodeError as e:
            raise ValueError(f"corrections JSON のパースに失敗しました: {e}")

        # ページ番号 → elements のマップを構築
        page_map: dict = {p["page_number"]: p["elements"] for p in layout}

        # コンテンツの有効範囲を計算（範囲外の corrections をクランプする）
        content_bounds: dict = {}
        for p in layout:
            pn = p["page_number"]
            border_elems = [e for e in p["elements"] if e.get("type") == "border_rect"]
            if border_elems:
                content_bounds[pn] = {
                    "max_row": max(e.get("end_row", e["row"]) for e in border_elems),
                    "max_col": max(e.get("end_col", e["col"]) for e in border_elems),
                }
            else:
                content_bounds[pn] = {"max_row": 9999, "max_col": 9999}

        applied = 0
        for c in corrections:
            action  = c.get("action")
            page_no = c.get("page", 1)
            elements = page_map.get(page_no)
            if elements is None:
                logger.warning(f"[correct] ページ {page_no} が見つかりません。スキップします。")
                continue

            if action == "add_text":
                elements.append({
                    "type": "text",
                    "content": c["content"],
                    "row": c["row"],
                    "col": c["col"],
                    "end_col": c["col"] + len(c["content"]),
                })
                applied += 1

            elif action == "fix_text":
                for elem in elements:
                    if elem.get("type") == "text" and elem["row"] == c["row"] and elem["col"] == c["col"]:
                        elem["row"] = c["new_row"]
                        elem["col"] = c["new_col"]
                        applied += 1
                        break

            elif action == "add_border":
                _end_row = c.get("end_row") or c.get("row_end", c["row"])
                _end_col = c.get("end_col") or c.get("col_end", c["col"])
                # コンテンツ範囲外の座標をクランプ
                bounds = content_bounds.get(page_no, {})
                _end_row = min(_end_row, bounds.get("max_row", _end_row))
                _end_col = min(_end_col, bounds.get("max_col", _end_col))
                elements.append({
                    "type": "border_rect",
                    "row": c["row"], "end_row": _end_row,
                    "col": c["col"], "end_col": _end_col,
                    "borders": c.get("borders", {"top": True, "bottom": True, "left": True, "right": True}),
                })
                applied += 1

            elif action == "remove_border":
                # 指定範囲に完全に包含される border_rect のみ削除する。
                # 重複（overlap）判定だと外枠など大きなボーダーが巻き添えで
                # 削除されてしまうため、包含（containment）判定を使用する。
                before = len(elements)
                _r  = c["row"]
                _er = c.get("end_row") or c.get("row_end", _r)
                _co = c["col"]
                _ec = c.get("end_col") or c.get("col_end", _co)
                elements[:] = [
                    e for e in elements
                    if not (e.get("type") == "border_rect"
                            and e["row"] >= _r and e["end_row"] <= _er
                            and e["col"] >= _co and e["end_col"] <= _ec)
                ]
                applied += before - len(elements)

        # 修正済みレイアウトを保存
        updated_json = json.dumps(layout, ensure_ascii=False)
        output_json_path.write_text(updated_json, encoding="utf-8")
        logger.info(f"[correct] {applied} 件の修正を適用しました: {output_json_path}")

    def rerender_after_corrections(
        self,
        pdf_name: str,
        grid_size: str,
        specific_out_dir: str = None,
    ) -> str:
        """
        correct コマンド用: 修正済み layout JSON + grid_params から Excel を再レンダリングする。

        auto_layout() が生成する grid_size サフィックス付きファイルを参照する。
          - {pdf_name}_{grid_size}_layout.json   （apply_corrections が更新済み）
          - {pdf_name}_{grid_size}_grid_params.json
        出力:
          - {pdf_name}_Python版.xlsx          （1pt/2pt 以外）
          - {pdf_name}_Python版_{grid_size}.xlsx  （1pt/2pt）
        """
        logger.info(f"--- [correct/rerender] Excel 再生成: {pdf_name} ({grid_size}) ---")
        out_dir = Path(specific_out_dir) if specific_out_dir else self.output_base_dir / pdf_name

        layout_path     = out_dir / f"{pdf_name}_{grid_size}_layout.json"
        grid_params_path = out_dir / f"{pdf_name}_{grid_size}_grid_params.json"
        _xlsx_suffix = f"_{grid_size}" if grid_size in ("1pt", "2pt") else ""
        xlsx_path    = out_dir / f"{pdf_name}_Python版{_xlsx_suffix}.xlsx"

        if not layout_path.exists():
            raise FileNotFoundError(f"layout JSON が見つかりません: {layout_path}")
        if not grid_params_path.exists():
            raise FileNotFoundError(f"grid_params JSON が見つかりません: {grid_params_path}")

        layout      = json.loads(layout_path.read_text(encoding="utf-8"))
        grid_params = json.loads(grid_params_path.read_text(encoding="utf-8"))

        _render_layout_to_xlsx(layout, grid_params, str(xlsx_path))
        logger.info(f"✅ correct/rerender 完了: {xlsx_path.name}")
        return str(xlsx_path)

