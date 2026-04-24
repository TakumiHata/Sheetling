import os
import tempfile
import json
import pytest
from src.renderer.excel import (
    render_layout_to_xlsx, fix_empty_cell_type_attr, _fit_row_height,
)


def _grid_params(**overrides):
    defaults = {
        'max_rows': 45, 'max_cols': 53,
        'excel_col_width': 1.74, 'excel_row_height': 18.25,
        'paper_size': 9, 'orientation': 'portrait',
        'default_font_size': 7, 'font_name': 'MS 明朝',
        'margin_left': 0.43, 'margin_right': 0.43,
        'margin_top': 0.41, 'margin_bottom': 0.41,
    }
    defaults.update(overrides)
    return defaults


class TestRenderLayoutToXlsx:
    def test_creates_file(self, tmp_path):
        layout = [{'page_number': 1, 'elements': []}]
        output = str(tmp_path / 'test.xlsx')
        render_layout_to_xlsx(layout, _grid_params(), output)
        assert os.path.exists(output)
        assert os.path.getsize(output) > 0

    def test_text_element(self, tmp_path):
        from openpyxl import load_workbook
        layout = [{
            'page_number': 1,
            'elements': [{'type': 'text', 'content': 'Hello', 'row': 3, 'col': 5}]
        }]
        output = str(tmp_path / 'test.xlsx')
        render_layout_to_xlsx(layout, _grid_params(), output)
        wb = load_workbook(output)
        ws = wb.active
        # row=3 + ROW_PADDING(1) = 4, col=5 + COL_OFFSET(1) = 6
        assert ws.cell(row=4, column=6).value == 'Hello'

    def test_border_element(self, tmp_path):
        from openpyxl import load_workbook
        layout = [{
            'page_number': 1,
            'elements': [{
                'type': 'border_rect', 'row': 2, 'end_row': 5, 'col': 3, 'end_col': 8,
                'borders': {'top': True, 'bottom': True, 'left': True, 'right': True},
            }]
        }]
        output = str(tmp_path / 'test.xlsx')
        render_layout_to_xlsx(layout, _grid_params(), output)
        wb = load_workbook(output)
        ws = wb.active
        # top-left cell: row=2+1=3, col=3+1=4
        cell = ws.cell(row=3, column=4)
        assert cell.border.top.style is not None

    def test_multi_page(self, tmp_path):
        layout = [
            {'page_number': 1, 'elements': [
                {'type': 'text', 'content': 'Page1', 'row': 1, 'col': 1}
            ]},
            {'page_number': 2, 'elements': [
                {'type': 'text', 'content': 'Page2', 'row': 1, 'col': 1}
            ]},
        ]
        output = str(tmp_path / 'test.xlsx')
        render_layout_to_xlsx(layout, _grid_params(), output)
        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb.active
        assert ws.cell(row=2, column=2).value == 'Page1'
        # page 2: row = 1 + (2-1)*(45+1) + 1 = 48
        assert ws.cell(row=48, column=2).value == 'Page2'

    def test_font_color(self, tmp_path):
        from openpyxl import load_workbook
        layout = [{
            'page_number': 1,
            'elements': [{
                'type': 'text', 'content': 'Red', 'row': 1, 'col': 1,
                'font_color': 'FF0000',
            }]
        }]
        output = str(tmp_path / 'test.xlsx')
        render_layout_to_xlsx(layout, _grid_params(), output)
        wb = load_workbook(output)
        ws = wb.active
        cell = ws.cell(row=2, column=2)
        assert 'FF0000' in str(cell.font.color.rgb)


class TestFitRowHeight:
    def test_sparse_keeps_default(self):
        # A4縦, 印刷可能 ~783pt: row=20 + padding → 21 * 18.25 = 383pt → 余裕で収まる
        layout = [{'page_number': 1, 'elements': [
            {'type': 'text', 'content': 'x', 'row': 20, 'col': 1}
        ]}]
        assert _fit_row_height(layout, _grid_params()) == 18.25

    def test_dense_shrinks(self):
        # A4縦 max_rows=45 をフルに使う → 46 * 18.25 = 839pt > 783pt → 17 に縮小
        # 46 * 17 = 782pt ≤ 783pt → 17 が選ばれる
        layout = [{'page_number': 1, 'elements': [
            {'type': 'text', 'content': 'bottom', 'row': 45, 'col': 1}
        ]}]
        assert _fit_row_height(layout, _grid_params()) == 17.0

    def test_border_end_row_exclusive(self):
        # end_row は排他的境界 → 実占有行は end_row - 1
        layout = [{'page_number': 1, 'elements': [
            {'type': 'border_rect', 'row': 1, 'end_row': 46, 'col': 1, 'end_col': 2}
        ]}]
        assert _fit_row_height(layout, _grid_params()) == 17.0

    def test_empty_layout(self):
        assert _fit_row_height([{'page_number': 1, 'elements': []}], _grid_params()) == 18.25

    def test_multi_page_uses_peak(self):
        # 複数ページのうち最大値で判定
        layout = [
            {'page_number': 1, 'elements': [{'type': 'text', 'content': 'a', 'row': 10, 'col': 1}]},
            {'page_number': 2, 'elements': [{'type': 'text', 'content': 'b', 'row': 45, 'col': 1}]},
        ]
        assert _fit_row_height(layout, _grid_params()) == 17.0


class TestFixEmptyCellTypeAttr:
    def test_removes_t_n(self, tmp_path):
        from openpyxl import Workbook
        from openpyxl.styles import Border, Side
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        cell.border = Border(top=Side(style='thin'))
        output = str(tmp_path / 'test.xlsx')
        wb.save(output)
        fix_empty_cell_type_attr(output)
        import zipfile
        with zipfile.ZipFile(output, 'r') as z:
            for name in z.namelist():
                if name.startswith('xl/worksheets/') and name.endswith('.xml'):
                    content = z.read(name).decode('utf-8')
                    assert 't="n"' not in content
