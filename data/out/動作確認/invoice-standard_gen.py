import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.worksheet.pagebreak import Break

# 入力データ（Step 1のレイアウト仕様）
data = [
    {
        "page_number": 1,
        "elements": [
            {
                "type": "text",
                "content": "請求書",
                "row": 5,
                "col": 14,
                "end_col": 21,
                "font_size": 20.0
            },
            {
                "type": "text",
                "content": "下記のとおりご請求いたします。",
                "row": 11,
                "col": 4,
                "end_col": 20,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "（振込先）",
                "row": 13,
                "col": 5,
                "end_col": 11,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "TEL.",
                "row": 13,
                "col": 21,
                "end_col": 23,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "◯◯銀行◯◯支店",
                "row": 14,
                "col": 5,
                "end_col": 14,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "FAX.",
                "row": 14,
                "col": 21,
                "end_col": 23,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "預金種別：",
                "row": 14,
                "col": 5,
                "end_col": 11,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "普通",
                "row": 14,
                "col": 10,
                "end_col": 12,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "口座番号：",
                "row": 15,
                "col": 5,
                "end_col": 10,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "1234567",
                "row": 15,
                "col": 10,
                "end_col": 17,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "口座名義：",
                "row": 16,
                "col": 5,
                "end_col": 10,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "◯◯株式会社",
                "row": 16,
                "col": 10,
                "end_col": 19,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "御請求金額",
                "row": 18,
                "col": 5,
                "end_col": 11,
                "font_size": 14.0
            },
            {
                "type": "text",
                "content": "¥32,400-",
                "row": 18,
                "col": 14,
                "end_col": 22,
                "font_size": 18.0
            },
            {
                "type": "text",
                "content": "（消費税込み）",
                "row": 18,
                "col": 21,
                "end_col": 28,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "年",
                "row": 7,
                "col": 27,
                "end_col": 28,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "月",
                "row": 7,
                "col": 30,
                "end_col": 31,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "日",
                "row": 7,
                "col": 33,
                "end_col": 34,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "品",
                "row": 20,
                "col": 10,
                "end_col": 10,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "名",
                "row": 20,
                "col": 10,
                "end_col": 10,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "数量",
                "row": 20,
                "col": 16,
                "end_col": 17,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "単価",
                "row": 20,
                "col": 21,
                "end_col": 22,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "金額",
                "row": 20,
                "col": 24,
                "end_col": 25,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "摘要",
                "row": 20,
                "col": 29,
                "end_col": 30,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "洗面台パイプ修理",
                "row": 21,
                "col": 4,
                "end_col": 14,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "1",
                "row": 21,
                "col": 16,
                "end_col": 16,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "21,000",
                "row": 21,
                "col": 21,
                "end_col": 23,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "21,000",
                "row": 21,
                "col": 24,
                "end_col": 26,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "サンプル",
                "row": 21,
                "col": 27,
                "end_col": 30,
                "font_size": 9.0
            },
            {
                "type": "text",
                "content": "清掃一式",
                "row": 23,
                "col": 4,
                "end_col": 9,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "1",
                "row": 23,
                "col": 16,
                "end_col": 16,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "9,000",
                "row": 23,
                "col": 21,
                "end_col": 23,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "9,000",
                "row": 23,
                "col": 24,
                "end_col": 26,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "サンプル",
                "row": 23,
                "col": 27,
                "end_col": 30,
                "font_size": 9.0
            },
            {
                "type": "text",
                "content": "消費税等",
                "row": 40,
                "col": 4,
                "end_col": 9,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "2,400",
                "row": 40,
                "col": 24,
                "end_col": 26,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "小",
                "row": 39,
                "col": 12,
                "end_col": 12,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "計",
                "row": 39,
                "col": 14,
                "end_col": 15,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "30,000",
                "row": 39,
                "col": 24,
                "end_col": 27,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "合",
                "row": 41,
                "col": 12,
                "end_col": 12,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "計",
                "row": 41,
                "col": 14,
                "end_col": 15,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "32,400",
                "row": 41,
                "col": 24,
                "end_col": 27,
                "font_size": 11.0
            },
            {
                "type": "text",
                "content": "備考：",
                "row": 43,
                "col": 4,
                "end_col": 7,
                "font_size": 11.0
            },
            {
                "type": "border_rect",
                "row": 15,
                "end_row": 16,
                "col": 5,
                "end_col": 10
            },
            {
                "type": "border_rect",
                "row": 15,
                "end_row": 16,
                "col": 26,
                "end_col": 30
            },
            {
                "type": "border_rect",
                "row": 15,
                "end_row": 16,
                "col": 30,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 16,
                "end_row": 19,
                "col": 5,
                "end_col": 10
            },
            {
                "type": "border_rect",
                "row": 16,
                "end_row": 19,
                "col": 26,
                "end_col": 30
            },
            {
                "type": "border_rect",
                "row": 16,
                "end_row": 19,
                "col": 30,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 20,
                "end_row": 21,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 20,
                "end_row": 21,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 20,
                "end_row": 21,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 20,
                "end_row": 21,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 20,
                "end_row": 21,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 21,
                "end_row": 23,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 21,
                "end_row": 23,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 21,
                "end_row": 23,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 21,
                "end_row": 23,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 21,
                "end_row": 23,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 23,
                "end_row": 24,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 23,
                "end_row": 24,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 23,
                "end_row": 24,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 23,
                "end_row": 24,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 23,
                "end_row": 24,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 24,
                "end_row": 25,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 24,
                "end_row": 25,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 24,
                "end_row": 25,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 24,
                "end_row": 25,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 24,
                "end_row": 25,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 25,
                "end_row": 27,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 25,
                "end_row": 27,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 25,
                "end_row": 27,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 25,
                "end_row": 27,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 25,
                "end_row": 27,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 27,
                "end_row": 28,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 27,
                "end_row": 28,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 27,
                "end_row": 28,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 27,
                "end_row": 28,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 27,
                "end_row": 28,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 28,
                "end_row": 29,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 28,
                "end_row": 29,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 28,
                "end_row": 29,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 28,
                "end_row": 29,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 28,
                "end_row": 29,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 29,
                "end_row": 31,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 29,
                "end_row": 31,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 29,
                "end_row": 31,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 29,
                "end_row": 31,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 29,
                "end_row": 31,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 31,
                "end_row": 32,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 31,
                "end_row": 32,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 31,
                "end_row": 32,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 31,
                "end_row": 32,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 31,
                "end_row": 32,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 32,
                "end_row": 33,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 32,
                "end_row": 33,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 32,
                "end_row": 33,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 32,
                "end_row": 33,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 32,
                "end_row": 33,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 33,
                "end_row": 35,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 33,
                "end_row": 35,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 33,
                "end_row": 35,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 33,
                "end_row": 35,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 33,
                "end_row": 35,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 35,
                "end_row": 36,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 35,
                "end_row": 36,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 35,
                "end_row": 36,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 35,
                "end_row": 36,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 35,
                "end_row": 36,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 36,
                "end_row": 37,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 36,
                "end_row": 37,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 36,
                "end_row": 37,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 36,
                "end_row": 37,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 36,
                "end_row": 37,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 37,
                "end_row": 39,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 37,
                "end_row": 39,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 37,
                "end_row": 39,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 37,
                "end_row": 39,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 37,
                "end_row": 39,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 39,
                "end_row": 40,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 39,
                "end_row": 40,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 39,
                "end_row": 40,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 39,
                "end_row": 40,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 39,
                "end_row": 40,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 40,
                "end_row": 41,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 40,
                "end_row": 41,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 40,
                "end_row": 41,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 40,
                "end_row": 41,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 40,
                "end_row": 41,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 41,
                "end_row": 43,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 41,
                "end_row": 43,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 41,
                "end_row": 43,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 41,
                "end_row": 43,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 41,
                "end_row": 43,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 43,
                "end_row": 47,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 43,
                "end_row": 47,
                "col": 16,
                "end_col": 18
            },
            {
                "type": "border_rect",
                "row": 43,
                "end_row": 47,
                "col": 18,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 43,
                "end_row": 47,
                "col": 24,
                "end_col": 27
            },
            {
                "type": "border_rect",
                "row": 43,
                "end_row": 47,
                "col": 27,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 5,
                "end_row": 5,
                "col": 14,
                "end_col": 21
            },
            {
                "type": "border_rect",
                "row": 9,
                "end_row": 9,
                "col": 14,
                "end_col": 14
            },
            {
                "type": "border_rect",
                "row": 9,
                "end_row": 9,
                "col": 21,
                "end_col": 26
            },
            {
                "type": "border_rect",
                "row": 10,
                "end_row": 10,
                "col": 21,
                "end_col": 33
            },
            {
                "type": "border_rect",
                "row": 11,
                "end_row": 11,
                "col": 21,
                "end_col": 26
            },
            {
                "type": "border_rect",
                "row": 10,
                "end_row": 10,
                "col": 4,
                "end_col": 16
            },
            {
                "type": "border_rect",
                "row": 19,
                "end_row": 19,
                "col": 5,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 6,
                "end_row": 6,
                "col": 14,
                "end_col": 24
            },
            {
                "type": "border_rect",
                "row": 6,
                "end_row": 6,
                "col": 14,
                "end_col": 24
            }
        ]
    }
]

# Excelファイルを作成
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

# グリッド設定 (列幅・行高)
for c in range(1, 37):
    ws.column_dimensions[get_column_letter(c)].width = 2.53
for r in range(1, 50 * len(data) + 1): # 仮に最大50行/ページとする
    ws.row_dimensions[r].height = 17.01

# 印刷設定
ws.page_setup.paperSize = 9
ws.page_setup.orientation = 'portrait'
ws.page_margins.left = 0.47
ws.page_margins.right = 0.47
ws.page_margins.top = 0.41
ws.page_margins.bottom = 0.41

# thin side for borders
thin = Side(style='thin')

def apply_outer_border(ws, s_row, e_row, s_col, e_col, fill_color=None):
    fill = PatternFill(patternType='solid', fgColor=fill_color) if fill_color else None
    for r in range(s_row, e_row + 1):
        for c in range(s_col, e_col + 1):
            target = ws.cell(row=r, column=c)
            try:
                target.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            except AttributeError:
                pass
            top    = thin if r == s_row else None
            bottom = thin if r == e_row else None
            left   = thin if c == s_col else None
            right  = thin if c == e_col else None
            try:
                target.border = Border(top=top, bottom=bottom, left=left, right=right)
            except AttributeError:
                pass
            if fill:
                try:
                    target.fill = fill
                except AttributeError:
                    pass

max_row_in_data = 0
max_col_in_data = 0

for page_data in data:
    row_offset = (page_data["page_number"] - 1) * 50
    
    for item in page_data["elements"]:
        if item["type"] == "text":
            r = item["row"] + row_offset
            c = item["col"]
            max_row_in_data = max(max_row_in_data, r)
            max_col_in_data = max(max_col_in_data, c)
            
            cell = ws.cell(row=r, column=c)
            cell.value = item["content"]
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            font_kwargs = {}
            if item.get("font_color"):
                font_kwargs["color"] = item["font_color"]
            if item.get("font_size"):
                font_kwargs["size"] = item["font_size"]
            if font_kwargs:
                cell.font = Font(**font_kwargs)
        
        elif item["type"] == "border_rect":
            s_row = item["row"] + row_offset
            e_row = item["end_row"] + row_offset
            s_col = item["col"]
            e_col = item["end_col"]
            max_row_in_data = max(max_row_in_data, e_row)
            max_col_in_data = max(max_col_in_data, e_col)
            
            apply_outer_border(
                ws,
                s_row, e_row,
                s_col, e_col,
                fill_color=item.get("fill_color")
            )
            
    # Add page break after each page's content (except the last page)
    if page_data["page_number"] < len(data):
        ws.row_breaks.append(Break(id=row_offset + 50)) # Add break at the start of the next page

# Set print range based on actual content
if max_row_in_data > 0 and max_col_in_data > 0:
    ws.print_area = f"A1:{get_column_letter(max_col_in_data)}{max_row_in_data}"

# Save the Excel file
wb.save("output.xlsx")