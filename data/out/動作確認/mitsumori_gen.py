import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

def create_estimate_excel(data_json):
    wb = openpyxl.Workbook()
    
    # JSONデータがリスト形式であることを想定
    for page_info in data_json:
        page_num = page_info.get("page_number", 1)
        if page_num == 1:
            ws = wb.active
            ws.title = f"Page {page_num}"
        else:
            ws = wb.create_sheet(title=f"Page {page_num}")

        # 1. レイアウトと方眼紙サイズの厳密な設定 [cite: 230]
        # 列幅設定 (1列＝約1.8mm) [cite: 230]
        for col in range(1, 101):
            ws.column_dimensions[get_column_letter(col)].width = 0.95
        
        # 行の高さ設定 (1行＝約5.3mm) [cite: 230]
        for row in range(1, 61):
            ws.row_dimensions[row].height = 15

        # 印刷設定 [cite: 230]
        ws.page_setup.paperSize = 9  # A4 [cite: 230]
        ws.page_setup.orientation = 'portrait'  # 縦向き [cite: 230]
        if "print_range" in page_info:
            ws.print_area = page_info["print_range"] # [cite: 230]

        # スタイル定義
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        header_font = Font(bold=True)
        default_alignment = Alignment(wrap_text=True, vertical='center')
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        merged_ranges = []

        # 4. 技術的制約の遵守 (セルの値を設定してから結合) [cite: 230]
        for item in page_info.get("data", []):
            s_row, s_col = item["start_row"], item["start_column"]
            e_row, e_col = item["end_row"], item["end_column"]
            val = item.get("value", "")
            has_border = item.get("border", False)

            # 重複チェックの簡易実装 [cite: 231]
            current_range = (s_row, s_col, e_row, e_col)
            overlap = False
            for r in merged_ranges:
                if not (e_row < r[0] or s_row > r[2] or e_col < r[1] or s_col > r[3]):
                    overlap = True
                    break
            if overlap:
                continue
            merged_ranges.append(current_range)

            # 起点セルに値を設定
            cell = ws.cell(row=s_row, column=s_col)
            cell.value = val
            cell.alignment = default_alignment

            # デザインの適用 [cite: 230]
            # 特定のキーワードや役割でヘッダー色を塗る
            if val in ["No.", "摘　　要", "数量", "標準価格", "見積価格", "合計金額", "承認", "担当営業"]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            # 結合処理 [cite: 230]
            if s_row != e_row or s_col != e_col:
                ws.merge_cells(start_row=s_row, start_column=s_col, end_row=e_row, end_column=e_col)

            # 罫線処理: 結合範囲全体に適用 [cite: 230]
            if has_border:
                for r in range(s_row, e_row + 1):
                    for c in range(s_col, e_col + 1):
                        ws.cell(row=r, column=c).border = thin_border

    # 5. 出力ファイル [cite: 230]
    wb.save("output.xlsx")

# 入力データ (STEP 5の出力) [cite: 231, 232, 233, 234, 235, 236, 237, 238, 239, 240, 241, 242, 243, 244, 245, 246, 247, 248, 249, 250, 251, 252, 253, 254, 255, 256, 257, 258, 259, 260, 261, 262]
input_data = [
  {
    "page_number": 1,
    "width": 595.0,
    "height": 842.0,
    "print_range": "B2:CV53",
    "data": [
      {"action": "merge_and_set", "start_row": 2, "start_column": 40, "end_row": 4, "end_column": 60, "value": "御見積書", "border": False},
      {"action": "merge_and_set", "start_row": 5, "start_column": 70, "end_row": 5, "end_column": 100, "value": "見積書Ｎｏ．20121119_00001", "border": False},
      {"action": "merge_and_set", "start_row": 6, "start_column": 5, "end_row": 7, "end_column": 45, "value": "△△株式会社 御中", "border": False},
      {"action": "merge_and_set", "start_row": 7, "start_column": 70, "end_row": 7, "end_column": 100, "value": "作成日: 2012年11月19日", "border": False},
      {"action": "merge_and_set", "start_row": 9, "start_column": 10, "end_row": 10, "end_column": 60, "value": "下記のとおり御見積申し上げます。\n何卒ご用命の程、お願い申し上げます。", "border": False},
      {"action": "merge_and_set", "start_row": 12, "start_column": 5, "end_row": 12, "end_column": 35, "value": "受渡期日:別途御打合せ", "border": False},
      {"action": "merge_and_set", "start_row": 12, "start_column": 65, "end_row": 16, "end_column": 100, "value": "ワークフロー商事株式会社\n東京都新宿区千代田９−９−９\nWSビル\nTEL  ０３-××××−９９９９\nFAX  ０３-９９９９−××××", "border": False},
      {"action": "merge_and_set", "start_row": 13, "start_column": 5, "end_row": 13, "end_column": 35, "value": "取引方法:別途御打合せ", "border": False},
      {"action": "merge_and_set", "start_row": 14, "start_column": 5, "end_row": 14, "end_column": 35, "value": "有効期限:発行日から30日", "border": False},
      {"action": "merge_and_set", "start_row": 15, "start_column": 5, "end_row": 15, "end_column": 35, "value": "貴社管理番号：", "border": False},
      {"action": "merge_and_set", "start_row": 15, "start_column": 80, "end_row": 15, "end_column": 90, "value": "承認", "border": True},
      {"action": "merge_and_set", "start_row": 15, "start_column": 91, "end_row": 15, "end_column": 100, "value": "担当営業", "border": True},
      {"action": "merge_and_set", "start_row": 16, "start_column": 80, "end_row": 18, "end_column": 90, "value": "", "border": True},
      {"action": "merge_and_set", "start_row": 16, "start_column": 91, "end_row": 18, "end_column": 100, "value": "", "border": True},
      {"action": "merge_and_set", "start_row": 20, "start_column": 20, "end_row": 21, "end_column": 80, "value": "合計金額： \\3,700,000\n(消費税別)", "border": False},
      {"action": "merge_and_set", "start_row": 23, "start_column": 2, "end_row": 23, "end_column": 7, "value": "No.", "border": True},
      {"action": "merge_and_set", "start_row": 23, "start_column": 8, "end_row": 23, "end_column": 50, "value": "摘　　要", "border": True},
      {"action": "merge_and_set", "start_row": 23, "start_column": 51, "end_row": 23, "end_column": 60, "value": "数量", "border": True},
      {"action": "merge_and_set", "start_row": 23, "start_column": 61, "end_row": 23, "end_column": 75, "value": "標準価格", "border": True},
      {"action": "merge_and_set", "start_row": 23, "start_column": 76, "end_row": 23, "end_column": 88, "value": "見積価格", "border": True},
      {"action": "merge_and_set", "start_row": 23, "start_column": 89, "end_row": 23, "end_column": 100, "value": "合計金額", "border": True},
      {"action": "merge_and_set", "start_row": 24, "start_column": 2, "end_row": 24, "end_column": 7, "value": "1", "border": True},
      {"action": "merge_and_set", "start_row": 24, "start_column": 8, "end_row": 24, "end_column": 50, "value": "ワークフローシステム　30ユーザーライセンス", "border": True},
      {"action": "merge_and_set", "start_row": 24, "start_column": 51, "end_row": 24, "end_column": 60, "value": "1", "border": True},
      {"action": "merge_and_set", "start_row": 24, "start_column": 61, "end_row": 24, "end_column": 75, "value": "2,700,000", "border": True},
      {"action": "merge_and_set", "start_row": 24, "start_column": 76, "end_row": 24, "end_column": 88, "value": "2,700,000", "border": True},
      {"action": "merge_and_set", "start_row": 24, "start_column": 89, "end_row": 24, "end_column": 100, "value": "2,700,000", "border": True},
      {"action": "merge_and_set", "start_row": 25, "start_column": 2, "end_row": 25, "end_column": 7, "value": "2", "border": True},
      {"action": "merge_and_set", "start_row": 25, "start_column": 8, "end_row": 25, "end_column": 50, "value": "初期設定費用", "border": True},
      {"action": "merge_and_set", "start_row": 25, "start_column": 51, "end_row": 25, "end_column": 60, "value": "1", "border": True},
      {"action": "merge_and_set", "start_row": 25, "start_column": 61, "end_row": 25, "end_column": 75, "value": "500,000", "border": True},
      {"action": "merge_and_set", "start_row": 25, "start_column": 76, "end_row": 25, "end_column": 88, "value": "500,000", "border": True},
      {"action": "merge_and_set", "start_row": 25, "start_column": 89, "end_row": 25, "end_column": 100, "value": "500,000", "border": True},
      {"action": "merge_and_set", "start_row": 26, "start_column": 2, "end_row": 26, "end_column": 7, "value": "3", "border": True},
      {"action": "merge_and_set", "start_row": 26, "start_column": 8, "end_row": 26, "end_column": 50, "value": "管理者費用", "border": True},
      {"action": "merge_and_set", "start_row": 26, "start_column": 51, "end_row": 26, "end_column": 60, "value": "1", "border": True},
      {"action": "merge_and_set", "start_row": 26, "start_column": 61, "end_row": 26, "end_column": 75, "value": "500,000", "border": True},
      {"action": "merge_and_set", "start_row": 26, "start_column": 76, "end_row": 26, "end_column": 88, "value": "500,000", "border": True},
      {"action": "merge_and_set", "start_row": 26, "start_column": 89, "end_row": 26, "end_column": 100, "value": "500,000", "border": True},
      {"action": "merge_and_set", "start_row": 47, "start_column": 2, "end_row": 47, "end_column": 88, "value": "合　　計", "border": True},
      {"action": "merge_and_set", "start_row": 47, "start_column": 89, "end_row": 47, "end_column": 100, "value": "3,700,000", "border": True},
      {"action": "merge_and_set", "start_row": 50, "start_column": 5, "end_row": 53, "end_column": 100, "value": "備　　考\n・消費税は別途計上させていただきます。\n・製品の瑕疵、無償保証期間は御購入後3ヶ月間です。", "border": False}
    ]
  }
]

if __name__ == "__main__":
    create_estimate_excel(input_data)