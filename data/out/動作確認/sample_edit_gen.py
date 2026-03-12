import json
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.worksheet.pagebreak import Break

# 入力データ
input_data = [
    {
        "page_number": 1,
        "width": 595.27557,
        "height": 841.88977,
        "print_range": "B1:AJ46",
        "data": [
            { "action": "set_value", "start_row": 1, "start_column": 14, "end_row": 2, "end_column": 23, "value": "SAMPLE PDF", "border": False },
            { "action": "set_value", "start_row": 3, "start_column": 2, "end_row": 5, "end_column": 36, "value": "ソースネクストの「いきなりPDF」は、販売本数シェアNo.1。高性能・低価格で操作も簡単。PDF作成の常識を変えたロング\nセラー製品です。", "border": False },
            { "action": "set_value", "start_row": 6, "start_column": 2, "end_row": 6, "end_column": 10, "value": "", "border": True },
            { "action": "set_value", "start_row": 6, "start_column": 11, "end_row": 6, "end_column": 19, "value": "いきなりPDF／BASIC", "border": True },
            { "action": "set_value", "start_row": 6, "start_column": 20, "end_row": 6, "end_column": 28, "value": "いきなりPDF／", "border": True },
            { "action": "set_value", "start_row": 6, "start_column": 29, "end_row": 6, "end_column": 36, "value": "いきなりPDF／", "border": True },
            { "action": "set_value", "start_row": 7, "start_column": 2, "end_row": 7, "end_column": 10, "value": "", "border": True },
            { "action": "set_value", "start_row": 7, "start_column": 11, "end_row": 7, "end_column": 19, "value": "Edition Ver.2", "border": True },
            { "action": "set_value", "start_row": 7, "start_column": 20, "end_row": 7, "end_column": 28, "value": "STANDARD Edition Ver.2", "border": True },
            { "action": "set_value", "start_row": 7, "start_column": 29, "end_row": 7, "end_column": 36, "value": "COMPLETE Edition Ver.2", "border": True },
            { "action": "set_value", "start_row": 8, "start_column": 2, "end_row": 8, "end_column": 10, "value": "ひとことで言うと", "border": True },
            { "action": "set_value", "start_row": 8, "start_column": 11, "end_row": 8, "end_column": 19, "value": "PDF作成・編集", "border": True },
            { "action": "set_value", "start_row": 8, "start_column": 20, "end_row": 8, "end_column": 28, "value": "PDF作成・データ変換・編 集", "border": True },
            { "action": "set_value", "start_row": 8, "start_column": 29, "end_row": 8, "end_column": 36, "value": "PDF作成・データ変換・高 度編集", "border": True },
            { "action": "set_value", "start_row": 9, "start_column": 2, "end_row": 9, "end_column": 10, "value": "標準価格（税込）", "border": True },
            { "action": "set_value", "start_row": 9, "start_column": 11, "end_row": 9, "end_column": 19, "value": "2,980円", "border": True },
            { "action": "set_value", "start_row": 9, "start_column": 20, "end_row": 9, "end_column": 28, "value": "3,980円", "border": True },
            { "action": "set_value", "start_row": 9, "start_column": 29, "end_row": 9, "end_column": 36, "value": "9,980円", "border": True },
            { "action": "set_value", "start_row": 10, "start_column": 2, "end_row": 10, "end_column": 10, "value": "Windows 8対応", "border": True },
            { "action": "set_value", "start_row": 10, "start_column": 11, "end_row": 10, "end_column": 19, "value": "○", "border": True },
            { "action": "set_value", "start_row": 10, "start_column": 20, "end_row": 10, "end_column": 28, "value": "○", "border": True },
            { "action": "set_value", "start_row": 10, "start_column": 29, "end_row": 10, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 13, "start_column": 2, "end_row": 13, "end_column": 36, "value": "PDFの作成", "border": False },
            { "action": "set_value", "start_row": 14, "start_column": 2, "end_row": 14, "end_column": 25, "value": "PDFファイルの作成", "border": True },
            { "action": "set_value", "start_row": 14, "start_column": 26, "end_row": 14, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 14, "start_column": 30, "end_row": 14, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 14, "start_column": 34, "end_row": 14, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 15, "start_column": 2, "end_row": 15, "end_column": 25, "value": "PDFファイルの閲覧、検索", "border": True },
            { "action": "set_value", "start_row": 15, "start_column": 26, "end_row": 15, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 15, "start_column": 30, "end_row": 15, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 15, "start_column": 34, "end_row": 15, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 16, "start_column": 2, "end_row": 16, "end_column": 25, "value": "Office製品へのプラグイン", "border": True },
            { "action": "set_value", "start_row": 16, "start_column": 26, "end_row": 16, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 16, "start_column": 30, "end_row": 16, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 16, "start_column": 34, "end_row": 16, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 17, "start_column": 2, "end_row": 17, "end_column": 25, "value": "フォントの埋め込み", "border": True },
            { "action": "set_value", "start_row": 17, "start_column": 26, "end_row": 17, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 17, "start_column": 30, "end_row": 17, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 17, "start_column": 34, "end_row": 17, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 18, "start_column": 2, "end_row": 18, "end_column": 25, "value": "複数文書の一括作成", "border": True },
            { "action": "set_value", "start_row": 18, "start_column": 26, "end_row": 18, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 18, "start_column": 30, "end_row": 18, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 18, "start_column": 34, "end_row": 18, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 20, "start_column": 2, "end_row": 20, "end_column": 36, "value": "PDFの組み換え", "border": False },
            { "action": "set_value", "start_row": 21, "start_column": 2, "end_row": 21, "end_column": 25, "value": "ページの分割、抽出、結合", "border": True },
            { "action": "set_value", "start_row": 21, "start_column": 26, "end_row": 21, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 21, "start_column": 30, "end_row": 21, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 21, "start_column": 34, "end_row": 21, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 23, "start_column": 2, "end_row": 23, "end_column": 36, "value": "PDF of 編集", "border": False },
            { "action": "set_value", "start_row": 24, "start_column": 2, "end_row": 24, "end_column": 25, "value": "ノート注釈の追加", "border": True },
            { "action": "set_value", "start_row": 24, "start_column": 26, "end_row": 24, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 24, "start_column": 30, "end_row": 24, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 24, "start_column": 34, "end_row": 24, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 25, "start_column": 2, "end_row": 25, "end_column": 25, "value": "テキストボックスの追加", "border": True },
            { "action": "set_value", "start_row": 25, "start_column": 26, "end_row": 25, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 25, "start_column": 30, "end_row": 25, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 25, "start_column": 34, "end_row": 25, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 26, "start_column": 2, "end_row": 26, "end_column": 25, "value": "添付ファイルの追加", "border": True },
            { "action": "set_value", "start_row": 26, "start_column": 26, "end_row": 26, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 26, "start_column": 30, "end_row": 26, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 26, "start_column": 34, "end_row": 26, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 27, "start_column": 2, "end_row": 27, "end_column": 25, "value": "しおりの作成、編集", "border": True },
            { "action": "set_value", "start_row": 27, "start_column": 26, "end_row": 27, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 27, "start_column": 30, "end_row": 27, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 27, "start_column": 34, "end_row": 27, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 28, "start_column": 2, "end_row": 28, "end_column": 25, "value": "ハイパーリンクの挿入", "border": True },
            { "action": "set_value", "start_row": 28, "start_column": 26, "end_row": 28, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 28, "start_column": 30, "end_row": 28, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 28, "start_column": 34, "end_row": 28, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 30, "start_column": 2, "end_row": 30, "end_column": 36, "value": "PDFを変換", "border": False },
            { "action": "set_value", "start_row": 31, "start_column": 2, "end_row": 31, "end_column": 25, "value": "PDFをWordに変換", "border": True },
            { "action": "set_value", "start_row": 31, "start_column": 26, "end_row": 31, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 31, "start_column": 30, "end_row": 31, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 31, "start_column": 34, "end_row": 31, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 32, "start_column": 2, "end_row": 32, "end_column": 25, "value": "PDFをExcelに変換", "border": True },
            { "action": "set_value", "start_row": 32, "start_column": 26, "end_row": 32, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 32, "start_column": 30, "end_row": 32, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 32, "start_column": 34, "end_row": 32, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 33, "start_column": 2, "end_row": 33, "end_column": 25, "value": "PDFをPowePointに変換", "border": True },
            { "action": "set_value", "start_row": 33, "start_column": 26, "end_row": 33, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 33, "start_column": 30, "end_row": 33, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 33, "start_column": 34, "end_row": 33, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 34, "start_column": 2, "end_row": 34, "end_column": 25, "value": "PDFをJPEGに変換", "border": True },
            { "action": "set_value", "start_row": 34, "start_column": 26, "end_row": 34, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 34, "start_column": 30, "end_row": 34, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 34, "start_column": 34, "end_row": 34, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 35, "start_column": 2, "end_row": 35, "end_column": 25, "value": "PDFをBMPに変換", "border": True },
            { "action": "set_value", "start_row": 35, "start_column": 26, "end_row": 35, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 35, "start_column": 30, "end_row": 35, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 35, "start_column": 34, "end_row": 35, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 36, "start_column": 2, "end_row": 36, "end_column": 25, "value": "透明テキスト付きPDFに変換", "border": True },
            { "action": "set_value", "start_row": 36, "start_column": 26, "end_row": 36, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 36, "start_column": 30, "end_row": 36, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 36, "start_column": 34, "end_row": 36, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 38, "start_column": 2, "end_row": 38, "end_column": 36, "value": "PDFの直接編集", "border": False },
            { "action": "set_value", "start_row": 39, "start_column": 2, "end_row": 39, "end_column": 25, "value": "すかしの挿入", "border": True },
            { "action": "set_value", "start_row": 39, "start_column": 26, "end_row": 39, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 39, "start_column": 30, "end_row": 39, "end_column": 33, "value": "×", "border": True },
            { "action": "set_value", "start_row": 39, "start_column": 34, "end_row": 39, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 40, "start_column": 2, "end_row": 40, "end_column": 25, "value": "クリップアートの挿入", "border": True },
            { "action": "set_value", "start_row": 40, "start_column": 26, "end_row": 40, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 40, "start_column": 30, "end_row": 40, "end_column": 33, "value": "×", "border": True },
            { "action": "set_value", "start_row": 40, "start_column": 34, "end_row": 40, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 41, "start_column": 2, "end_row": 41, "end_column": 25, "value": "スタンプの追加", "border": True },
            { "action": "set_value", "start_row": 41, "start_column": 26, "end_row": 41, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 41, "start_column": 30, "end_row": 41, "end_column": 33, "value": "×", "border": True },
            { "action": "set_value", "start_row": 41, "start_column": 34, "end_row": 41, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 42, "start_column": 2, "end_row": 42, "end_column": 25, "value": "ページのトリミング編集", "border": True },
            { "action": "set_value", "start_row": 42, "start_column": 26, "end_row": 42, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 42, "start_column": 30, "end_row": 42, "end_column": 33, "value": "×", "border": True },
            { "action": "set_value", "start_row": 42, "start_column": 34, "end_row": 42, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 43, "start_column": 2, "end_row": 43, "end_column": 25, "value": "フォームオブジェクトの追加", "border": True },
            { "action": "set_value", "start_row": 43, "start_column": 26, "end_row": 43, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 43, "start_column": 30, "end_row": 43, "end_column": 33, "value": "×", "border": True },
            { "action": "set_value", "start_row": 43, "start_column": 34, "end_row": 43, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 44, "start_column": 2, "end_row": 44, "end_column": 25, "value": "テキストの直接編集", "border": True },
            { "action": "set_value", "start_row": 44, "start_column": 26, "end_row": 44, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 44, "start_column": 30, "end_row": 44, "end_column": 33, "value": "×", "border": True },
            { "action": "set_value", "start_row": 44, "start_column": 34, "end_row": 44, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 45, "start_column": 2, "end_row": 45, "end_column": 25, "value": "オブジェクトの編集", "border": True },
            { "action": "set_value", "start_row": 45, "start_column": 26, "end_row": 45, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 45, "start_column": 30, "end_row": 45, "end_column": 33, "value": "×", "border": True },
            { "action": "set_value", "start_row": 45, "start_column": 34, "end_row": 45, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 46, "start_column": 2, "end_row": 46, "end_column": 25, "value": "ページの回転編集", "border": True },
            { "action": "set_value", "start_row": 46, "start_column": 26, "end_row": 46, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 46, "start_column": 30, "end_row": 46, "end_column": 33, "value": "×", "border": True },
            { "action": "set_value", "start_row": 46, "start_column": 34, "end_row": 46, "end_column": 36, "value": "○", "border": True }
        ]
    },
    {
        "page_number": 2,
        "width": 595.27557,
        "height": 841.88977,
        "print_range": "B1:AJ7",
        "data": [
            { "action": "set_value", "start_row": 1, "start_column": 2, "end_row": 1, "end_column": 36, "value": "セキュリティ", "border": False },
            { "action": "set_value", "start_row": 2, "start_column": 2, "end_row": 2, "end_column": 25, "value": "暗号化", "border": True },
            { "action": "set_value", "start_row": 2, "start_column": 26, "end_row": 2, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 2, "start_column": 30, "end_row": 2, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 2, "start_column": 34, "end_row": 2, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 3, "start_column": 2, "end_row": 3, "end_column": 25, "value": "閲覧制限", "border": True },
            { "action": "set_value", "start_row": 3, "start_column": 26, "end_row": 3, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 3, "start_column": 30, "end_row": 3, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 3, "start_column": 34, "end_row": 3, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 4, "start_column": 2, "end_row": 4, "end_column": 25, "value": "印刷制限", "border": True },
            { "action": "set_value", "start_row": 4, "start_column": 26, "end_row": 4, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 4, "start_column": 30, "end_row": 4, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 4, "start_column": 34, "end_row": 4, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 5, "start_column": 2, "end_row": 5, "end_column": 25, "value": "修正制限", "border": True },
            { "action": "set_value", "start_row": 5, "start_column": 26, "end_row": 5, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 5, "start_column": 30, "end_row": 5, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 5, "start_column": 34, "end_row": 5, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 6, "start_column": 2, "end_row": 6, "end_column": 25, "value": "コピー制限", "border": True },
            { "action": "set_value", "start_row": 6, "start_column": 26, "end_row": 6, "end_column": 29, "value": "○", "border": True },
            { "action": "set_value", "start_row": 6, "start_column": 30, "end_row": 6, "end_column": 33, "value": "○", "border": True },
            { "action": "set_value", "start_row": 6, "start_column": 34, "end_row": 6, "end_column": 36, "value": "○", "border": True },
            { "action": "set_value", "start_row": 7, "start_column": 2, "end_row": 7, "end_column": 25, "value": "電子署名の添付", "border": True },
            { "action": "set_value", "start_row": 7, "start_column": 26, "end_row": 7, "end_column": 29, "value": "×", "border": True },
            { "action": "set_value", "start_row": 7, "start_column": 30, "end_row": 7, "end_column": 33, "value": "×", "border": True },
            { "action": "set_value", "start_row": 7, "start_column": 34, "end_row": 7, "end_column": 36, "value": "○", "border": True }
        ]
    }
]

wb = Workbook()
ws = wb.active

# 1. レイアウトと方眼紙サイズの厳密な設定
num_pages = len(input_data)
max_rows = num_pages * 50

for i in range(1, 37):
    ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 2.53

for i in range(1, max_rows + 1):
    ws.row_dimensions[i].height = 17.01

# デザイン用スタイル定義
thin = Side(border_style="thin", color="000000")
border_style = Border(top=thin, left=thin, right=thin, bottom=thin)
header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
header_font = Font(bold=True)
default_alignment = Alignment(wrap_text=True, vertical='center')

# 4. 複数ページへの完全対応
for page in input_data:
    page_num = page["page_number"]
    offset = (page_num - 1) * 50
    
    for item in page["data"]:
        s_row = item["start_row"] + offset
        e_row = item["end_row"] + offset
        s_col = item["start_column"]
        e_col = item["end_column"]
        val = item["value"]
        is_border = item.get("border", False)
        
        # 5. 技術的制約の遵守 (openpyxl)
        try:
            # 値の設定は左上のセルのみに行う
            main_cell = ws.cell(row=s_row, column=s_col)
            main_cell.value = val
            
            # ヘッダー推定 (簡易的にBorderありの1行目付近や特定の文字を太字にするなどの処理)
            if is_border and s_row - offset <= 7:
                main_cell.fill = header_fill
                main_cell.font = header_font

            # 範囲全体に対して二重ループで枠線や折り返し設定を適用
            for r in range(s_row, e_row + 1):
                for c in range(s_col, e_col + 1):
                    target_cell = ws.cell(row=r, column=c)
                    target_cell.alignment = default_alignment
                    if is_border:
                        target_cell.border = border_style
        except AttributeError:
            pass

    # 改ページ設定
    if page_num < num_pages:
        ws.row_breaks.append(Break(id=offset + 50))

# 3. 印刷設定
ws.page_setup.paperSize = 9
ws.page_setup.orientation = 'portrait'
ws.print_options.horizontalCentered = True
ws.page_margins.left = 0.47
ws.page_margins.right = 0.47
ws.page_margins.top = 0.41
ws.page_margins.bottom = 0.41

# 印刷範囲の動的設定 (全ページ分)
last_page = input_data[-1]
last_offset = (last_page["page_number"] - 1) * 50
ws.print_area = f"A1:AJ{last_offset + 50}"

wb.save("output.xlsx")