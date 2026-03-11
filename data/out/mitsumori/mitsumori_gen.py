import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

def create_excel_from_commands(data_list):
    wb = openpyxl.Workbook()
    # デフォルトのシートを削除して各ページを新規作成
    wb.remove(wb.active)

    for page_data in data_list:
        page_num = page_data.get("page_number", 1)
        ws = wb.create_sheet(title=f"Page {page_num}")
        
        # 1. レイアウトと方眼紙サイズの厳密な設定
        # 最大100列まで設定
        for col in range(1, 101):
            ws.column_dimensions[get_column_letter(col)].width = 0.95
        # 最大55行まで設定
        for row in range(1, 56):
            ws.row_dimensions[row].height = 15

        # スタイルの定義
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        header_font = Font(bold=True)

        commands = page_data.get("data", [])
        
        # 結合範囲の重複チェック用
        merged_ranges = []

        for cmd in commands:
            sr, sc = cmd["start_row"], cmd["start_column"]
            er, ec = cmd["end_row"], cmd["end_column"]
            val = cmd.get("value", "")
            has_border = cmd.get("border", False)

            # 技術的制約：起点セルに値を設定
            cell = ws.cell(row=sr, column=sc)
            cell.value = val

            # 基本アライメント（折り返し、中央揃え）
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

            # 特殊なデザインの適用 (ヘッダー等)
            # 内容に基づいて簡易的に判定、またはSTEP3/4の属性を利用可能だが、ここでは"摘 要"などを含む行をヘッダーとする
            if val in ["No.", "摘 要", "数量", "標準価格", "見積価格", "合計金額", "承認", "担当営業", "備 考"]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 見積書のタイトルは大きく中央揃え
            if val == "御見積書":
                cell.font = Font(size=16, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # セル結合の実行
            if sr != er or sc != ec:
                new_range = f"{get_column_letter(sc)}{sr}:{get_column_letter(ec)}{er}"
                
                # 重複チェック
                overlap = False
                for r in merged_ranges:
                    if new_range == r: # 簡易チェック
                        overlap = True
                        break
                
                if not overlap:
                    ws.merge_cells(new_range)
                    merged_ranges.append(new_range)

            # 罫線の適用 (結合セル全体に適用するためにループ処理)
            if has_border:
                for r in range(sr, er + 1):
                    for c in range(sc, ec + 1):
                        ws.cell(row=r, column=c).border = thin_border

        # 3. 印刷設定
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        
        # 印刷範囲の設定
        if "print_range" in page_data:
            ws.print_area = page_data["print_range"]

        # スケーリング設定は要件により記述しない（等倍100%）

    # 5. 保存
    wb.save("output.xlsx")

# 入力データ
input_data = [
  {
    "page_number": 1,
    "width": 595.0,
    "height": 842.0,
    "print_range": "B3:CV53",
    "data": [
      {"action": "merge_and_set", "start_row": 3, "start_column": 40, "end_row": 5, "end_column": 60, "value": "御見積書", "border": False},
      {"action": "merge_and_set", "start_row": 6, "start_column": 70, "end_row": 6, "end_column": 95, "value": "見積書Ｎｏ．20121119_00001", "border": False},
      {"action": "merge_and_set", "start_row": 7, "start_column": 5, "end_row": 8, "end_column": 45, "value": "△△株式会社 御中", "border": False},
      {"action": "merge_and_set", "start_row": 8, "start_column": 70, "end_row": 8, "end_column": 95, "value": "作成日: 2012年11月19日", "border": False},
      {"action": "merge_and_set", "start_row": 10, "start_column": 10, "end_row": 11, "end_column": 50, "value": "下記のとおり御見積申し上げます。\n何卒ご用命の程、お願い申し上げます。", "border": False},
      {"action": "merge_and_set", "start_row": 13, "start_column": 5, "end_row": 13, "end_column": 30, "value": "受渡期日:別途御打合せ", "border": False},
      {"action": "merge_and_set", "start_row": 13, "start_column": 70, "end_row": 13, "end_column": 95, "value": "ワークフロー商事株式会社", "border": False},
      {"action": "merge_and_set", "start_row": 14, "start_column": 5, "end_row": 14, "end_column": 30, "value": "取引方法:別途御打合せ", "border": False},
      {"action": "merge_and_set", "start_row": 14, "start_column": 70, "end_row": 15, "end_column": 95, "value": "東京都新宿区千代田９−９−９\nWSビル", "border": False},
      {"action": "merge_and_set", "start_row": 16, "start_column": 5, "end_row": 16, "end_column": 30, "value": "有効期限:発行日から30日", "border": False},
      {"action": "merge_and_set", "start_row": 16, "start_column": 70, "end_row": 17, "end_column": 95, "value": "TEL ０３-××××−９９９９\nFAX ０３-９９９９−××××", "border": False},
      {"action": "merge_and_set", "start_row": 18, "start_column": 5, "end_row": 18, "end_column": 20, "value": "貴社管理番号：", "border": False},
      {"action": "merge_and_set", "start_row": 18, "start_column": 75, "end_row": 18, "end_column": 84, "value": "承認", "border": True},
      {"action": "merge_and_set", "start_row": 18, "start_column": 85, "end_row": 18, "end_column": 95, "value": "担当営業", "border": True},
      {"action": "merge_and_set", "start_row": 19, "start_column": 75, "end_row": 21, "end_column": 84, "value": "", "border": True},
      {"action": "merge_and_set", "start_row": 19, "start_column": 85, "end_row": 21, "end_column": 95, "value": "", "border": True},
      {"action": "merge_and_set", "start_row": 22, "start_column": 20, "end_row": 23, "end_column": 70, "value": "合計金額： \\3,700,000 (消費税別)", "border": False},
      {"action": "merge_and_set", "start_row": 25, "start_column": 2, "end_row": 25, "end_column": 10, "value": "No.", "border": True},
      {"action": "merge_and_set", "start_row": 25, "start_column": 11, "end_row": 25, "end_column": 55, "value": "摘 要", "border": True},
      {"action": "merge_and_set", "start_row": 25, "start_column": 56, "end_row": 25, "end_column": 65, "value": "数量", "border": True},
      {"action": "merge_and_set", "start_row": 25, "start_column": 66, "end_row": 25, "end_column": 77, "value": "標準価格", "border": True},
      {"action": "merge_and_set", "start_row": 25, "start_column": 78, "end_row": 25, "end_column": 89, "value": "見積価格", "border": True},
      {"action": "merge_and_set", "start_row": 25, "start_column": 90, "end_row": 25, "end_column": 100, "value": "合計金額", "border": True},
      {"action": "merge_and_set", "start_row": 26, "start_column": 2, "end_row": 26, "end_column": 10, "value": "1", "border": True},
      {"action": "merge_and_set", "start_row": 26, "start_column": 11, "end_row": 26, "end_column": 55, "value": "ワークフローシステム 30ユーザーライセンス", "border": True},
      {"action": "merge_and_set", "start_row": 26, "start_column": 56, "end_row": 26, "end_column": 65, "value": "1", "border": True},
      {"action": "merge_and_set", "start_row": 26, "start_column": 66, "end_row": 26, "end_column": 77, "value": "2,700,000", "border": True},
      {"action": "merge_and_set", "start_row": 26, "start_column": 78, "end_row": 26, "end_column": 89, "value": "2,700,000", "border": True},
      {"action": "merge_and_set", "start_row": 26, "start_column": 90, "end_row": 26, "end_column": 100, "value": "2,700,000", "border": True},
      {"action": "merge_and_set", "start_row": 27, "start_column": 2, "end_row": 27, "end_column": 10, "value": "2", "border": True},
      {"action": "merge_and_set", "start_row": 27, "start_column": 11, "end_row": 27, "end_column": 55, "value": "初期設定費用", "border": True},
      {"action": "merge_and_set", "start_row": 27, "start_column": 56, "end_row": 27, "end_column": 65, "value": "1", "border": True},
      {"action": "merge_and_set", "start_row": 27, "start_column": 66, "end_row": 27, "end_column": 77, "value": "500,000", "border": True},
      {"action": "merge_and_set", "start_row": 27, "start_column": 78, "end_row": 27, "end_column": 89, "value": "500,000", "border": True},
      {"action": "merge_and_set", "start_row": 27, "start_column": 90, "end_row": 27, "end_column": 100, "value": "500,000", "border": True},
      {"action": "merge_and_set", "start_row": 28, "start_column": 2, "end_row": 28, "end_column": 10, "value": "3", "border": True},
      {"action": "merge_and_set", "start_row": 28, "start_column": 11, "end_row": 28, "end_column": 55, "value": "管理者費用", "border": True},
      {"action": "merge_and_set", "start_row": 28, "start_column": 56, "end_row": 28, "end_column": 65, "value": "1", "border": True},
      {"action": "merge_and_set", "start_row": 28, "start_column": 66, "end_row": 28, "end_column": 77, "value": "500,000", "border": True},
      {"action": "merge_and_set", "start_row": 28, "start_column": 78, "end_row": 28, "end_column": 89, "value": "500,000", "border": True},
      {"action": "merge_and_set", "start_row": 28, "start_column": 90, "end_row": 28, "end_column": 100, "value": "500,000", "border": True},
      {"action": "merge_and_set", "start_row": 49, "start_column": 2, "end_row": 49, "end_column": 89, "value": "合 計", "border": True},
      {"action": "merge_and_set", "start_row": 49, "start_column": 90, "end_row": 49, "end_column": 100, "value": "3,700,000", "border": True},
      {"action": "merge_and_set", "start_row": 51, "start_column": 2, "end_row": 51, "end_column": 100, "value": "備 考", "border": True},
      {"action": "merge_and_set", "start_row": 52, "start_column": 2, "end_row": 52, "end_column": 100, "value": "・消費税は別途計上させていただきます。", "border": True},
      {"action": "merge_and_set", "start_row": 53, "start_column": 2, "end_row": 53, "end_column": 100, "value": "・製品の瑕疵、無償保証期間は御購入後3ヶ月間です。", "border": True}
    ]
  }
]

if __name__ == "__main__":
    create_excel_from_commands(input_data)