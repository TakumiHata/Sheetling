import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

MAX_ROWS = 76
COL_OFFSET = 1
ROW_PADDING = 2
EXCEL_COL_WIDTH = 1.45
EXCEL_ROW_HEIGHT = 11.34
DEFAULT_FONT_SIZE = 7

_dir = Path(__file__).parent
data = json.loads((_dir / "prompts" / "001598405_step1_5_output.json").read_text(encoding="utf-8"))

wb = Workbook()
ws = wb.active
thin = Side(style='thin')
total_pages = len(data)

for c in range(1, 200):
    ws.column_dimensions[get_column_letter(c)].width = EXCEL_COL_WIDTH
for r in range(1, MAX_ROWS * total_pages + ROW_PADDING + 1):
    ws.row_dimensions[r].height = EXCEL_ROW_HEIGHT


def apply_border(ws, s_row, e_row, s_col, e_col, borders):
    has_top    = borders.get("top",    True)
    has_bottom = borders.get("bottom", True)
    has_left   = borders.get("left",   True)
    has_right  = borders.get("right",  True)
    for r in range(s_row, e_row + 1):
        for c in range(s_col, e_col + 1):
            target = ws.cell(row=r, column=c)
            try:
                target.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            except AttributeError:
                pass
            top    = thin if (r == s_row and has_top)    else None
            bottom = thin if (r == e_row and has_bottom) else None
            left   = thin if (c == s_col and has_left)   else None
            right  = thin if (c == e_col and has_right)  else None
            try:
                target.border = Border(top=top, bottom=bottom, left=left, right=right)
            except AttributeError:
                pass


max_used_row = ROW_PADDING
max_used_col = COL_OFFSET

for page in data:
    page_number = page["page_number"]
    row_offset = (page_number - 1) * MAX_ROWS + ROW_PADDING

    if page_number > 1:
        ws.row_breaks.append(Break(id=(page_number - 1) * MAX_ROWS + ROW_PADDING))

    for item in page["elements"]:
        if item["type"] == "text":
            r = item["row"] + row_offset
            c = item["col"] + COL_OFFSET
            try:
                cell = ws.cell(row=r, column=c)
                cell.value = item["content"]
                if item.get("is_vertical"):
                    cell.alignment = Alignment(text_rotation=255, vertical='top', wrap_text=False)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
                font_kwargs = {"size": item.get("font_size") or DEFAULT_FONT_SIZE}
                if item.get("font_color"):
                    font_kwargs["color"] = item["font_color"]
                cell.font = Font(**font_kwargs)
                max_used_row = max(max_used_row, r)
                max_used_col = max(max_used_col, c)
            except AttributeError:
                pass

        elif item["type"] == "border_rect":
            s_row = item["row"] + row_offset
            e_row = item["end_row"] + row_offset
            s_col = item["col"] + COL_OFFSET
            e_col = item["end_col"] + COL_OFFSET
            apply_border(ws, s_row, e_row, s_col, e_col,
                         item.get("borders", {"top": True, "bottom": True, "left": True, "right": True}))
            max_used_row = max(max_used_row, e_row)
            max_used_col = max(max_used_col, e_col)

ws.page_setup.paperSize = 9
ws.page_setup.orientation = 'portrait'
ws.page_margins.left = 0.43
ws.page_margins.right = 0.43
ws.page_margins.top = 0.41
ws.page_margins.bottom = 0.41
ws.print_area = f"A1:{get_column_letter(max_used_col)}{max_used_row}"

wb.save("output.xlsx")
