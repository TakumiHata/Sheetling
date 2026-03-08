import math
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generate(wb, ws):
    """
    PDFから抽出されたテキスト・座標・スタイル情報を基に、
    方眼Excel（120列×176行/ページ）上にレイアウトを再現します。
    """
    import math

    # テキスト要素の全件リスト
    elements = [
        {"text": "請求書", "x0": 58.87, "top": 60.45, "x1": 163.27, "bottom": 95.25, "fontname": "Noto-Sans-JP-Thin", "size": 34.8, "page_width": 960.0, "color": "224466"},
        {"text": "株式会社クライアント", "x0": 58.87, "top": 127.5, "x1": 275.5, "bottom": 149.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": " ", "x0": 275.5, "top": 132.96, "x1": 281.48, "bottom": 154.71, "fontname": "SegoeUI-Semibold", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "様", "x0": 281.48, "top": 127.5, "x1": 303.23, "bottom": 149.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "発⾏⽇", "x0": 58.87, "top": 172.5, "x1": 124.12, "bottom": 194.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": ": 2026", "x0": 124.12, "top": 177.96, "x1": 181.7, "bottom": 199.71, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "年", "x0": 181.7, "top": 172.5, "x1": 203.45, "bottom": 194.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "3", "x0": 203.45, "top": 177.96, "x1": 215.17, "bottom": 199.71, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "⽉", "x0": 215.17, "top": 172.5, "x1": 236.92, "bottom": 194.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "8", "x0": 236.92, "top": 177.96, "x1": 248.65, "bottom": 199.71, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "⽇", "x0": 248.65, "top": 172.5, "x1": 270.4, "bottom": 194.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "請求番号", "x0": 58.87, "top": 204.75, "x1": 145.87, "bottom": 226.5, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": ": INV-20260308-01", "x0": 145.87, "top": 210.21, "x1": 326.76, "bottom": 231.96, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "平素は格別のお引き⽴てを賜り、厚く御礼申し上げます。", "x0": 58.87, "top": 249.75, "x1": 620.89, "bottom": 271.5, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "下記の通り、ご請求申し上げます。", "x0": 58.87, "top": 282.0, "x1": 403.83, "bottom": 303.75, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "ご請求⾦額", "x0": 58.87, "top": 330.97, "x1": 200.25, "bottom": 359.25, "fontname": "Noto-Sans-JP-Thin", "size": 28.27, "page_width": 960.0, "color": "1F2328"},
        {"text": "¥ 165,000 (", "x0": 58.87, "top": 401.43, "x1": 227.98, "bottom": 436.23, "fontname": "SegoeUI-Semibold", "size": 34.8, "page_width": 960.0, "color": "224466"},
        {"text": "税込", "x0": 227.98, "top": 392.7, "x1": 297.58, "bottom": 427.5, "fontname": "Noto-Sans-JP-Thin", "size": 34.8, "page_width": 960.0, "color": "224466"},
        {"text": ")", "x0": 297.58, "top": 401.43, "x1": 309.14, "bottom": 436.23, "fontname": "SegoeUI-Semibold", "size": 34.8, "page_width": 960.0, "color": "224466"},
        {"text": "※明細およびお振込先につきましては、次ページ以降をご参照ください。", "x0": 58.87, "top": 460.5, "x1": 775.1, "bottom": 482.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "請求書（サンプル）", "x0": 22.5, "top": 17.25, "x1": 143.6, "bottom": 30.75, "fontname": "Noto-Sans-JP-Thin", "size": 13.5, "page_width": 960.0, "color": "666666"},
        {"text": "株式会社サンプル", "x0": 22.5, "top": 505.5, "x1": 130.1, "bottom": 519.0, "fontname": "Noto-Sans-JP-Thin", "size": 13.5, "page_width": 960.0, "color": "666666"},
        {"text": "1", "x0": 927.8, "top": 504.77, "x1": 937.5, "bottom": 522.77, "fontname": "SegoeUI", "size": 18.0, "page_width": 960.0, "color": "777777"},
        {"text": "ご請求明細", "x0": 58.87, "top": 629.7, "x1": 232.87, "bottom": 664.5, "fontname": "Noto-Sans-JP-Thin", "size": 34.8, "page_width": 960.0, "color": "224466"},
        {"text": "No.", "x0": 69.37, "top": 708.21, "x1": 104.29, "bottom": 729.96, "fontname": "SegoeUI-Semibold", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "品名", "x0": 124.55, "top": 702.75, "x1": 168.05, "bottom": 724.5, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "数量", "x0": 436.28, "top": 702.75, "x1": 479.78, "bottom": 724.5, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "単価", "x0": 500.03, "top": 702.75, "x1": 543.53, "bottom": 724.5, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "⾦額", "x0": 601.3, "top": 702.75, "x1": 644.8, "bottom": 724.5, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "1", "x0": 69.37, "top": 750.21, "x1": 81.1, "bottom": 771.96, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "Web", "x0": 124.55, "top": 750.21, "x1": 168.17, "bottom": 771.96, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "サイトデザイン制作⼀式", "x0": 168.17, "top": 744.75, "x1": 407.42, "bottom": 766.5, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "1", "x0": 436.28, "top": 750.21, "x1": 448.0, "bottom": 771.96, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "¥ 50,000", "x0": 500.03, "top": 750.21, "x1": 581.05, "bottom": 771.96, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "¥ 50,000", "x0": 601.3, "top": 750.21, "x1": 682.32, "bottom": 771.96, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "2", "x0": 69.37, "top": 792.96, "x1": 81.1, "bottom": 814.71, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "フロントエンド実装（", "x0": 124.55, "top": 787.5, "x1": 342.05, "bottom": 809.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "React", "x0": 342.05, "top": 792.96, "x1": 394.28, "bottom": 814.71, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "）", "x0": 394.28, "top": 787.5, "x1": 416.03, "bottom": 809.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "1", "x0": 436.28, "top": 792.96, "x1": 448.0, "bottom": 814.71, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "¥ 70,000", "x0": 500.03, "top": 792.96, "x1": 581.05, "bottom": 814.71, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "¥ 70,000", "x0": 601.3, "top": 792.96, "x1": 682.32, "bottom": 814.71, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "3", "x0": 69.37, "top": 834.96, "x1": 81.1, "bottom": 856.71, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "バックエンド実装（", "x0": 124.55, "top": 829.5, "x1": 320.3, "bottom": 851.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "Python", "x0": 320.3, "top": 834.96, "x1": 387.73, "bottom": 856.71, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "）", "x0": 387.73, "top": 829.5, "x1": 409.48, "bottom": 851.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "1", "x0": 436.28, "top": 834.96, "x1": 448.0, "bottom": 856.71, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "¥ 30,000", "x0": 500.03, "top": 834.96, "x1": 581.05, "bottom": 856.71, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "¥ 30,000", "x0": 601.3, "top": 834.96, "x1": 682.32, "bottom": 856.71, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "⼩計", "x0": 124.55, "top": 872.25, "x1": 168.05, "bottom": 894.0, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "¥ 150,000", "x0": 601.3, "top": 877.71, "x1": 694.05, "bottom": 899.46, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "消費税", "x0": 124.55, "top": 914.25, "x1": 189.8, "bottom": 936.0, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": " (10%)", "x0": 189.8, "top": 919.71, "x1": 252.65, "bottom": 941.46, "fontname": "SegoeUI-Semibold", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "¥ 15,000", "x0": 601.3, "top": 919.71, "x1": 682.32, "bottom": 941.46, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "合計", "x0": 124.55, "top": 957.0, "x1": 168.05, "bottom": 978.75, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "¥ 165,000", "x0": 601.3, "top": 962.46, "x1": 697.05, "bottom": 984.21, "fontname": "SegoeUI-Semibold", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "請求書（サンプル）", "x0": 22.5, "top": 557.25, "x1": 143.6, "bottom": 570.75, "fontname": "Noto-Sans-JP-Thin", "size": 13.5, "page_width": 960.0, "color": "666666"},
        {"text": "株式会社サンプル", "x0": 22.5, "top": 1045.5, "x1": 130.1, "bottom": 1059.0, "fontname": "Noto-Sans-JP-Thin", "size": 13.5, "page_width": 960.0, "color": "666666"},
        {"text": "2", "x0": 927.8, "top": 1044.77, "x1": 937.5, "bottom": 1062.77, "fontname": "SegoeUI", "size": 18.0, "page_width": 960.0, "color": "777777"},
        {"text": "お振込先‧特記事項", "x0": 58.87, "top": 1140.45, "x1": 372.07, "bottom": 1175.25, "fontname": "Noto-Sans-JP-Thin", "size": 34.8, "page_width": 960.0, "color": "224466"},
        {"text": "銀⾏振込先", "x0": 58.87, "top": 1211.47, "x1": 200.25, "bottom": 1239.75, "fontname": "Noto-Sans-JP-Thin", "size": 28.27, "page_width": 960.0, "color": "1F2328"},
        {"text": "⾦融機関名", "x0": 102.37, "top": 1269.75, "x1": 211.12, "bottom": 1291.5, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": ": ", "x0": 211.12, "top": 1275.21, "x1": 221.8, "bottom": 1296.96, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "サンプル銀⾏", "x0": 221.8, "top": 1269.75, "x1": 351.65, "bottom": 1291.5, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "⽀店名", "x0": 102.37, "top": 1308.0, "x1": 167.62, "bottom": 1329.75, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": ": ", "x0": 167.62, "top": 1313.46, "x1": 178.3, "bottom": 1335.21, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "ビジネス⽀店", "x0": 178.3, "top": 1308.0, "x1": 308.8, "bottom": 1329.75, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": " (", "x0": 308.8, "top": 1313.46, "x1": 321.32, "bottom": 1335.21, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "店番", "x0": 321.32, "top": 1308.0, "x1": 364.82, "bottom": 1329.75, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": ": 123)", "x0": 364.82, "top": 1313.46, "x1": 417.23, "bottom": 1335.21, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "⼝座種別", "x0": 102.37, "top": 1345.5, "x1": 189.37, "bottom": 1367.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": ": ", "x0": 189.37, "top": 1350.96, "x1": 200.05, "bottom": 1372.71, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "普通⼝座", "x0": 200.05, "top": 1345.5, "x1": 287.05, "bottom": 1367.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "⼝座番号", "x0": 102.37, "top": 1383.75, "x1": 189.37, "bottom": 1405.5, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": ": 1234567", "x0": 189.37, "top": 1389.21, "x1": 282.12, "bottom": 1410.96, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "⼝座名義", "x0": 102.37, "top": 1422.0, "x1": 189.37, "bottom": 1443.75, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": ": ", "x0": 189.37, "top": 1427.46, "x1": 200.05, "bottom": 1449.21, "fontname": "SegoeUI", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "カ）サンプル", "x0": 200.05, "top": 1422.0, "x1": 329.9, "bottom": 1443.75, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "お⽀払い期限", "x0": 58.87, "top": 1470.22, "x1": 228.52, "bottom": 1498.5, "fontname": "Noto-Sans-JP-Thin", "size": 28.27, "page_width": 960.0, "color": "1F2328"},
        {"text": "2026", "x0": 58.87, "top": 1533.96, "x1": 107.24, "bottom": 1555.71, "fontname": "SegoeUI-Semibold", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "年", "x0": 107.24, "top": 1528.5, "x1": 128.99, "bottom": 1550.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "4", "x0": 128.99, "top": 1533.96, "x1": 141.52, "bottom": 1555.71, "fontname": "SegoeUI-Semibold", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "⽉", "x0": 141.52, "top": 1528.5, "x1": 163.27, "bottom": 1550.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "30", "x0": 163.27, "top": 1533.96, "x1": 187.42, "bottom": 1555.71, "fontname": "SegoeUI-Semibold", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "⽇", "x0": 187.42, "top": 1528.5, "x1": 209.17, "bottom": 1550.25, "fontname": "Noto-Sans-JP-Thin", "size": 21.75, "page_width": 960.0, "color": "1F2328"},
        {"text": "特記事項", "x0": 58.87, "top": 1576.72, "x1": 171.97, "bottom": 1605.0, "fontname": "Noto-Sans-JP-Thin", "size": 28.27, "page_width": 960.0, "color": "1F2328"},
        {"text": "請求書（サンプル）", "x0": 22.5, "top": 1097.25, "x1": 143.6, "bottom": 1110.75, "fontname": "Noto-Sans-JP-Thin", "size": 13.5, "page_width": 960.0, "color": "666666"},
        {"text": "株式会社サンプル", "x0": 22.5, "top": 1585.5, "x1": 130.1, "bottom": 1599.0, "fontname": "Noto-Sans-JP-Thin", "size": 13.5, "page_width": 960.0, "color": "666666"},
        {"text": "3", "x0": 927.8, "top": 1584.77, "x1": 937.5, "bottom": 1602.77, "fontname": "SegoeUI", "size": 18.0, "page_width": 960.0, "color": "777777"}
    ]

    # テキスト要素の処理
    for el in elements:
        # スケーリング計算 (1マス=4.65pt)
        sr = math.floor(el["top"] / 4.65) + 1
        sc = math.floor(el["x0"] / el["page_width"] * 120) + 1
        
        # セルへの書き込み
        cell = ws.cell(row=sr, column=sc)
        cell.value = el["text"]
        
        # フォント設定（色指定がある場合は#を除去して適用）
        font_color = el["color"].replace("#", "")
        cell.font = Font(name=el["fontname"], size=el["size"], color=font_color)
        
        # 配置設定
        cell.alignment = Alignment(vertical='center', wrap_text=False)

    # 矩形データ（罫線等）の処理
    rects = [
        {"x0": 59.25, "top": 694.5, "x1": 114.75, "bottom": 737.25},
        {"x0": 114.75, "top": 694.5, "x1": 426.0, "bottom": 737.25},
        {"x0": 426.0, "top": 694.5, "x1": 489.75, "bottom": 737.25},
        {"x0": 489.75, "top": 694.5, "x1": 591.0, "bottom": 737.25},
        {"x0": 591.0, "top": 694.5, "x1": 707.25, "bottom": 737.25},
        {"x0": 59.25, "top": 737.25, "x1": 114.75, "bottom": 779.25},
        {"x0": 114.75, "top": 737.25, "x1": 426.0, "bottom": 779.25},
        {"x0": 426.0, "top": 737.25, "x1": 489.75, "bottom": 779.25},
        {"x0": 489.75, "top": 737.25, "x1": 591.0, "bottom": 779.25},
        {"x0": 591.0, "top": 737.25, "x1": 707.25, "bottom": 779.25},
        {"x0": 59.25, "top": 779.25, "x1": 114.75, "bottom": 822.0},
        {"x0": 114.75, "top": 779.25, "x1": 426.0, "bottom": 822.0},
        {"x0": 426.0, "top": 779.25, "x1": 489.75, "bottom": 822.0},
        {"x0": 489.75, "top": 779.25, "x1": 591.0, "bottom": 822.0},
        {"x0": 591.0, "top": 779.25, "x1": 707.25, "bottom": 822.0},
        {"x0": 59.25, "top": 822.0, "x1": 114.75, "bottom": 864.0},
        {"x0": 114.75, "top": 822.0, "x1": 426.0, "bottom": 864.0},
        {"x0": 426.0, "top": 822.0, "x1": 489.75, "bottom": 864.0},
        {"x0": 489.75, "top": 822.0, "x1": 591.0, "bottom": 864.0},
        {"x0": 591.0, "top": 822.0, "x1": 707.25, "bottom": 864.0},
        {"x0": 59.25, "top": 864.0, "x1": 114.75, "bottom": 906.75},
        {"x0": 114.75, "top": 864.0, "x1": 426.0, "bottom": 906.75},
        {"x0": 426.0, "top": 864.0, "x1": 489.75, "bottom": 906.75},
        {"x0": 489.75, "top": 864.0, "x1": 591.0, "bottom": 906.75},
        {"x0": 591.0, "top": 864.0, "x1": 707.25, "bottom": 906.75},
        {"x0": 59.25, "top": 906.75, "x1": 114.75, "bottom": 948.75},
        {"x0": 114.75, "top": 906.75, "x1": 426.0, "bottom": 948.75},
        {"x0": 426.0, "top": 906.75, "x1": 489.75, "bottom": 948.75},
        {"x0": 489.75, "top": 906.75, "x1": 591.0, "bottom": 948.75},
        {"x0": 591.0, "top": 906.75, "x1": 707.25, "bottom": 948.75},
        {"x0": 59.25, "top": 948.75, "x1": 114.75, "bottom": 991.5},
        {"x0": 114.75, "top": 948.75, "x1": 426.0, "bottom": 991.5},
        {"x0": 426.0, "top": 948.75, "x1": 489.75, "bottom": 991.5},
        {"x0": 489.75, "top": 948.75, "x1": 591.0, "bottom": 991.5},
        {"x0": 591.0, "top": 948.75, "x1": 707.25, "bottom": 991.5}
    ]

    # 罫線スタイルの定義
    thin_side = Side(style='thin', color='000000')

    # 矩形データに基づいてセルの周囲に罫線を設定
    for r in rects:
        # ページ横幅は共通して960.0とする
        page_width = 960.0
        r_sr = math.floor(r["top"] / 4.65) + 1
        r_er = math.ceil(r["bottom"] / 4.65)
        r_sc = math.floor(r["x0"] / page_width * 120) + 1
        r_ec = math.ceil(r["x1"] / page_width * 120)

        for row_idx in range(r_sr, r_er + 1):
            for col_idx in range(r_sc, r_ec + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                # 簡易的に外周のみに薄い罫線を適用（実務上は結合なしのため各セルに設定）
                current_border = c.border
                new_border = Border(
                    left=thin_side if col_idx == r_sc else current_border.left,
                    right=thin_side if col_idx == r_ec else current_border.right,
                    top=thin_side if row_idx == r_sr else current_border.top,
                    bottom=thin_side if row_idx == r_er else current_border.bottom
                )
                c.border = new_border